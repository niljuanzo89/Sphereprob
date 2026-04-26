import streamlit as st
import csv
import math
import io
import os
import ssl
import json
import difflib
import random
import datetime
import textwrap
import urllib.request
import urllib.parse
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# reportlab imported lazily inside build_bingo_pdf

FILEPATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "phish_net_setlists_2016_2026.csv")
START_YEAR = 2008

TIER_TARGETS = {
    "Staple":     0.036,
    "Common":     0.390,
    "Occasional": 0.333,
    "Rare":       0.240,
}

def get_tier(pct):
    if pct >= 15: return "Staple"
    if pct >= 5:  return "Common"
    if pct >= 1:  return "Occasional"
    return "Rare"

def avg_position(pos_list):
    return sum(pos_list) / len(pos_list)

@st.cache_data
def load_data():
    global_counter = Counter()
    global_shows = 0
    city_data = {}
    song_last_show = {}
    song_positions = defaultdict(list)

    with open(FILEPATH, newline="") as f:
        rows = [r for r in csv.DictReader(f) if r["date"][:4].isdigit() and int(r["date"][:4]) >= START_YEAR]

    show_index = 0
    for row in rows:
        setlist = row["setlist"].strip()
        if not setlist:
            continue
        songs = [s.strip() for s in setlist.split("|") if s.strip()]
        if not songs:
            continue

        global_shows += 1
        global_counter.update(songs)
        total = len(songs)
        loc = row["location"].strip()

        if loc not in city_data:
            city_data[loc] = {"counter": Counter(), "shows": 0, "lengths": [], "positions": defaultdict(list)}
        city_data[loc]["counter"].update(songs)
        city_data[loc]["shows"] += 1
        city_data[loc]["lengths"].append(total)

        for i, song in enumerate(songs):
            norm_pos = i / max(total - 1, 1)
            song_positions[song].append(norm_pos)
            city_data[loc]["positions"][song].append(norm_pos)
            song_last_show[song] = show_index

        show_index += 1

    total_shows = show_index
    current_gap = {song: total_shows - last - 1 for song, last in song_last_show.items()}
    return global_counter, global_shows, city_data, current_gap, total_shows, song_positions


BUSTOUT_GAP = 500


def build_structured_setlist(scores, pos_fn, current_gap, total_shows):
    """Select songs for a 2-set + 2-encore show structure.

    Returns list of dicts: [{'song':..., 'set':'Set 1'|'Set 2'|'Encore', 'role':'Opener'|'Closer'|'E1'|'E2'|''}]
    in play order. Enforces:
      - Each set: 6-8 songs, exactly 1 opener + 1 closer
      - 2 encores (E1, E2)
      - At most 1 song with gap > 500 across the whole show (bust-out)
    Positions in data are 0-1 across the full show (both sets + encore concatenated).
    """
    set1_size = random.choice([6, 7, 8])
    set2_size = random.choice([6, 7, 8])

    # Songs sorted by score, descending
    all_cands = sorted(scores.keys(), key=lambda s: scores[s], reverse=True)

    # Positional pools (score-ordered)
    s1_opener_pool = [s for s in all_cands if pos_fn(s) < 0.20]
    s1_closer_pool = [s for s in all_cands if 0.30 <= pos_fn(s) <= 0.50]
    s2_opener_pool = [s for s in all_cands if 0.45 <= pos_fn(s) <= 0.65]
    s2_closer_pool = [s for s in all_cands if 0.70 <= pos_fn(s) <= 0.90]
    encore_pool    = [s for s in all_cands if pos_fn(s) > 0.85]
    s1_mid_pool    = [s for s in all_cands if 0.15 <= pos_fn(s) <= 0.45]
    s2_mid_pool    = [s for s in all_cands if 0.50 <= pos_fn(s) <= 0.85]

    used = set()
    bust_kept = False

    def pick(pool):
        nonlocal bust_kept
        for s in pool:
            if s in used:
                continue
            g = current_gap.get(s, total_shows)
            if g > BUSTOUT_GAP and bust_kept:
                continue
            used.add(s)
            if g > BUSTOUT_GAP:
                bust_kept = True
            return s
        return None

    def pick_fallback():
        return pick(all_cands)

    s1_open  = pick(s1_opener_pool) or pick_fallback()
    s1_close = pick(s1_closer_pool) or pick_fallback()
    s2_open  = pick(s2_opener_pool) or pick_fallback()
    s2_close = pick(s2_closer_pool) or pick_fallback()
    enc1     = pick(encore_pool)    or pick_fallback()
    enc2     = pick(encore_pool)    or pick_fallback()

    # Fill middles
    s1_mid = []
    for s in s1_mid_pool:
        if len(s1_mid) >= set1_size - 2:
            break
        if s in used:
            continue
        g = current_gap.get(s, total_shows)
        if g > BUSTOUT_GAP and bust_kept:
            continue
        used.add(s)
        if g > BUSTOUT_GAP:
            bust_kept = True
        s1_mid.append(s)

    s2_mid = []
    for s in s2_mid_pool:
        if len(s2_mid) >= set2_size - 2:
            break
        if s in used:
            continue
        g = current_gap.get(s, total_shows)
        if g > BUSTOUT_GAP and bust_kept:
            continue
        used.add(s)
        if g > BUSTOUT_GAP:
            bust_kept = True
        s2_mid.append(s)

    # Top off from any remaining if pools were too thin
    while len(s1_mid) < set1_size - 2:
        s = pick_fallback()
        if not s:
            break
        s1_mid.append(s)
    while len(s2_mid) < set2_size - 2:
        s = pick_fallback()
        if not s:
            break
        s2_mid.append(s)

    # Order middles by position
    s1_mid.sort(key=pos_fn)
    s2_mid.sort(key=pos_fn)

    out = []
    if s1_open:  out.append({"song": s1_open,  "set": "Set 1", "role": "Opener"})
    for s in s1_mid:
        out.append({"song": s, "set": "Set 1", "role": ""})
    if s1_close: out.append({"song": s1_close, "set": "Set 1", "role": "Closer"})
    if s2_open:  out.append({"song": s2_open,  "set": "Set 2", "role": "Opener"})
    for s in s2_mid:
        out.append({"song": s, "set": "Set 2", "role": ""})
    if s2_close: out.append({"song": s2_close, "set": "Set 2", "role": "Closer"})
    if enc1:     out.append({"song": enc1, "set": "Encore", "role": "E1"})
    if enc2:     out.append({"song": enc2, "set": "Encore", "role": "E2"})
    return out


def generate_setlist(city):
    global_counter, global_shows, city_data, current_gap, total_shows, song_positions = load_data()

    matches = {loc: v for loc, v in city_data.items() if city.lower() in loc.lower()}
    if not matches:
        return None, None, None

    city_counter = Counter()
    city_shows = 0
    all_lengths = []
    all_positions = defaultdict(list)
    matched_locations = []

    for loc, d in matches.items():
        city_counter.update(d["counter"])
        city_shows += d["shows"]
        all_lengths.extend(d["lengths"])
        matched_locations.append(loc)
        for song, positions in d["positions"].items():
            all_positions[song].extend(positions)

    avg_length = round(sum(all_lengths) / len(all_lengths))

    scores = {}
    for song, count in city_counter.items():
        freq = count / city_shows
        gap = current_gap.get(song, total_shows)
        gap_boost = 1 + math.log(gap + 1) / math.log(total_shows + 1)
        scores[song] = freq * gap_boost

    def song_pos(s):
        return avg_position(all_positions.get(s, song_positions.get(s, [0.5])))

    structured = build_structured_setlist(scores, song_pos, current_gap, total_shows)

    rows = []
    for i, entry in enumerate(structured, 1):
        song = entry["song"]
        base_pct = (city_counter[song] / city_shows) * 100
        gpct = (global_counter[song] / global_shows) * 100
        gap = current_gap.get(song, total_shows)
        adj = scores[song] * 100
        pos = song_pos(song)
        rows.append({
            "#": i,
            "Song": song,
            "Set": entry["set"],
            "Role": entry["role"],
            "Tier": get_tier(gpct),
            "City Freq": f"{base_pct:.1f}%",
            "Shows Since Last Played": gap,
            "Adj Score": f"{adj:.1f}%",
            "Show Position": entry["role"] if entry["role"] else f"{pos:.0%} thru",
            "Bust Out": gap > BUSTOUT_GAP,
            "_adj": adj,
            "_gap": gap,
            "_pos": pos,
        })

    return rows, city_shows, matched_locations


def generate_bingo(city):
    global_counter, global_shows, city_data, current_gap, total_shows, song_positions = load_data()
    rows, city_shows, locations = generate_setlist(city)
    if rows is None:
        return None

    def gap_score(song):
        gap = current_gap.get(song, total_shows)
        return (global_counter[song] / global_shows) * (1 + math.log(gap + 1) / math.log(total_shows + 1))

    setlist_picks = [r["Song"] for r in sorted(rows, key=lambda r: r["_adj"], reverse=True)[:10]]
    used = set(setlist_picks)

    common_pool = sorted(
        [(s, gap_score(s)) for s, c in global_counter.items()
         if 5 <= (c / global_shows) * 100 < 15 and s not in used],
        key=lambda x: x[1], reverse=True
    )
    common_picks = [s for s, _ in common_pool[:10]]
    used |= set(common_picks)

    occasional_pool = sorted(
        [(s, gap_score(s)) for s, c in global_counter.items()
         if 1 <= (c / global_shows) * 100 < 5 and s not in used],
        key=lambda x: x[1], reverse=True
    )
    rare_pool = sorted(
        [(s, gap_score(s)) for s, c in global_counter.items()
         if (c / global_shows) * 100 < 1 and s not in used],
        key=lambda x: x[1], reverse=True
    )
    rare_picks = [s for s, _ in occasional_pool[:3]] + [s for s, _ in rare_pool[:2]]

    all_songs = setlist_picks + common_picks + rare_picks
    random.shuffle(all_songs)

    setlist_set = set(setlist_picks)
    common_set  = set(common_picks)
    return [{"song": s,
             "cat": "setlist" if s in setlist_set else "common" if s in common_set else "rare"}
            for s in all_songs]


def generate_sphere_setlist(target_date, sphere_songs_played):
    """Predict a setlist for an upcoming Sphere show.

    Uses the same tier/gap/position logic as generate_setlist, but:
    - Based on Las Vegas city history (Sphere is in Las Vegas)
    - Excludes any song already played at the Sphere 2026 run so far
    - Adds a boost for songs the band played 10-15 days before target_date
      (captures "current rotation" songs)
    """
    global_counter, global_shows, city_data, current_gap, total_shows, song_positions = load_data()

    already_played = set(sphere_songs_played.keys())

    # Use Las Vegas data for the Sphere (fall back to global if sparse)
    matches = {loc: v for loc, v in city_data.items() if "las vegas" in loc.lower()}

    if matches:
        city_counter = Counter()
        city_shows_n = 0
        all_lengths = []
        all_positions = defaultdict(list)
        matched_locations = []
        for loc, d in matches.items():
            city_counter.update(d["counter"])
            city_shows_n += d["shows"]
            all_lengths.extend(d["lengths"])
            matched_locations.append(loc)
            for song, positions in d["positions"].items():
                all_positions[song].extend(positions)
        avg_length = round(sum(all_lengths) / len(all_lengths))
        source = f"Las Vegas ({city_shows_n} shows)"
    else:
        city_counter = global_counter
        city_shows_n = global_shows
        all_positions = song_positions
        avg_length = 22
        matched_locations = ["Global"]
        source = f"Global ({global_shows} shows)"

    # Find songs played in the 10-15 shows prior to target_date — "recent rotation" boost.
    # (Looks at the 10th through 15th most recent shows before target_date, chronologically.)
    all_shows = []
    with open(FILEPATH, newline="") as f:
        for row in csv.DictReader(f):
            if not row["date"][:4].isdigit():
                continue
            songs_here = [s.strip() for s in row["setlist"].split("|") if s.strip()]
            if songs_here:
                all_shows.append((row["date"], songs_here))
    all_shows.sort(key=lambda x: x[0])

    prior_shows = [s for s in all_shows if s[0] < target_date]
    # Window = shows indexed [-15 .. -10] from the end of prior_shows
    window_shows = prior_shows[-15:-9] if len(prior_shows) >= 10 else prior_shows[:max(0, len(prior_shows)-9)]

    recent_songs = set()
    recent_dates = []
    for date_str, songs_here in window_shows:
        recent_songs.update(songs_here)
        recent_dates.append(date_str)

    # Score songs (excluding already-played)
    scores = {}
    for song, count in city_counter.items():
        if song in already_played:
            continue
        freq = count / city_shows_n
        gap = current_gap.get(song, total_shows)
        gap_boost    = 1 + math.log(gap + 1) / math.log(total_shows + 1)
        recent_boost = 1.6 if song in recent_songs else 1.0
        scores[song] = freq * gap_boost * recent_boost

    def song_pos(s):
        return avg_position(all_positions.get(s, song_positions.get(s, [0.5])))

    structured = build_structured_setlist(scores, song_pos, current_gap, total_shows)

    rows = []
    for i, entry in enumerate(structured, 1):
        song = entry["song"]
        base_pct = (city_counter[song] / city_shows_n) * 100
        gpct     = (global_counter[song] / global_shows) * 100
        gap      = current_gap.get(song, total_shows)
        adj      = scores[song] * 100
        pos      = song_pos(song)
        rows.append({
            "#": i,
            "Song": song,
            "Set": entry["set"],
            "Role": entry["role"],
            "Tier": get_tier(gpct),
            "Vegas/Global Freq": f"{base_pct:.1f}%",
            "Global Freq": f"{gpct:.1f}%",
            "Shows Since Last Played": gap,
            "Adj Score": f"{adj:.1f}%",
            "Show Position": entry["role"] if entry["role"] else f"{pos:.0%} thru",
            "Recent": song in recent_songs,
            "Bust Out": gap > 500,
            "_adj": adj,
            "_gap": gap,
            "_pos": pos,
        })

    return {
        "rows": rows,
        "source": source,
        "city_shows": city_shows_n,
        "excluded": sorted(already_played),
        "recent_songs": sorted(recent_songs),
        "recent_dates": sorted(set(recent_dates)),
        "window_count": len(window_shows),
    }


def _inject_free_center(song_cards, size):
    """Take a list of song cards and return a size×size list with a FREE center cell.

    Accepts up to (size*size - 1) song cards; pads with duplicates if short.
    """
    needed = size * size - 1
    songs = list(song_cards)[:needed]
    # Pad if we somehow didn't get enough cards
    while len(songs) < needed:
        songs.append({"song": "—", "cat": "rare"})
    # 🥚 1-in-100 easter egg: FREE square swaps to FUEGO
    free_label = "🔥 FUEGO 🔥" if random.random() < 0.01 else "★ FREE ★"
    free_card = {"song": free_label, "cat": "setlist", "free": True}
    center = (size * size) // 2  # 4 for 3×3, 12 for 5×5
    return songs[:center] + [free_card] + songs[center:]


def build_bingo_pdf(cards, city, size=5):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.6*inch, rightMargin=0.6*inch,
        topMargin=0.5*inch, bottomMargin=0.5*inch,
    )
    CAT_BG = {
        "setlist": colors.HexColor("#2a2a10"),
        "common":  colors.HexColor("#0d2235"),
        "rare":    colors.HexColor("#2a0d35"),
    }
    CAT_FG = {
        "setlist": colors.HexColor("#F0E68C"),
        "common":  colors.HexColor("#7ec8e3"),
        "rare":    colors.HexColor("#ce93d8"),
    }
    BORDER    = colors.HexColor("#444444")
    HEADER_BG = colors.HexColor("#1a1a2e")
    HEADER_FG = colors.HexColor("#F0E68C")

    story = []
    story.append(Paragraph("🎸 Gotta-Jibbootistics", ParagraphStyle(
        "t", fontSize=20, textColor=HEADER_FG, alignment=TA_CENTER,
        fontName="Helvetica-Bold", spaceAfter=2)))
    story.append(Paragraph("whatever you do, take care of your shoes", ParagraphStyle(
        "s", fontSize=9, textColor=colors.HexColor("#e85545"), alignment=TA_CENTER,
        fontName="Helvetica-Oblique", spaceAfter=4)))
    story.append(Paragraph(f"Bingo Card — {city.title()}", ParagraphStyle(
        "c", fontSize=11, textColor=colors.HexColor("#aaaacc"), alignment=TA_CENTER,
        fontName="Helvetica", spaceAfter=10)))

    # Sizing: 5×5 uses tighter cells; 3×3 uses bigger cells
    if size == 3:
        col_w, row_h, hdr_h, font_size, leading = 2.1*inch, 1.5*inch, 0.45*inch, 12, 14
        header_chars = ["P", "H", "I"]
    else:
        col_w, row_h, hdr_h, font_size, leading = 1.26*inch, 0.85*inch, 0.38*inch, 7.5, 10
        header_chars = list("PHISH")

    header_row = [Paragraph(ch, ParagraphStyle(
        "h", fontSize=18, textColor=HEADER_FG, alignment=TA_CENTER,
        fontName="Helvetica-Bold")) for ch in header_chars]

    FREE_BG = colors.HexColor("#3a2800")
    FREE_FG = colors.HexColor("#FFF3B0")

    table_data = [header_row]
    for row_i in range(size):
        row = []
        for col_i in range(size):
            card = cards[row_i * size + col_i]
            fg = FREE_FG if card.get("free") else CAT_FG[card["cat"]]
            row.append(Paragraph(card["song"], ParagraphStyle(
                f"cell", fontSize=font_size, textColor=fg,
                alignment=TA_CENTER, fontName="Helvetica-Bold", leading=leading)))
        table_data.append(row)

    tbl = Table(table_data, colWidths=[col_w]*size, rowHeights=[hdr_h]+[row_h]*size)
    ts = [
        ("BACKGROUND", (0, 0), (size-1, 0), HEADER_BG),
        ("GRID",       (0, 0), (-1, -1), 1.5, BORDER),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
    ]
    for row_i in range(size):
        for col_i in range(size):
            card = cards[row_i*size + col_i]
            bg = FREE_BG if card.get("free") else CAT_BG[card["cat"]]
            ts.append(("BACKGROUND", (col_i, row_i+1), (col_i, row_i+1), bg))
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "🟡 From predicted setlist  |  🔵 Globally common  |  🟣 Rare / uncommon",
        ParagraphStyle("leg", fontSize=7.5, textColor=colors.HexColor("#888888"),
                       alignment=TA_CENTER, fontName="Helvetica")))
    doc.build(story)
    buf.seek(0)
    return buf


def build_xlsx(rows, city, city_shows):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{city} Setlist"

    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"PHISH  —  {city.upper()}"
    c.font = Font(name="Arial", bold=True, size=18, color="F0E68C")
    c.fill = PatternFill("solid", fgColor="1A1A2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Predicted Setlist  |  Based on {city_shows} shows  |  Gap-adjusted · Position-ordered"
    c.font = Font(name="Arial", italic=True, size=10, color="AAAACC")
    c.fill = PatternFill("solid", fgColor="1A1A2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    headers = ["#", "Song", "Tier", "City Freq\n(Base %)", "Shows Since\nLast Played", "Adj Score", "Show Position\n(Avg % thru)"]
    col_widths = [5, 32, 12, 14, 16, 12, 16]
    for col, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=col, value=hdr)
        c.font = Font(name="Arial", bold=True, size=10, color="F0E68C")
        c.fill = PatternFill("solid", fgColor="1A1A2E")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[4].height = 34

    for i, row in enumerate(rows):
        r = i + 5
        is_closer  = row["_pos"] > 0.85
        is_bustout = row.get("Bust Out", False)
        is_alt     = i % 2 == 1

        if is_closer:   bg, fg = "2E1A1A", "F4C2C2"
        elif is_bustout: bg, fg = "1A2E1A", "B9F6CA"
        else:           bg, fg = ("EAEAF5" if is_alt else "FFFFFF"), "111111"

        adj_color = ("1B5E20" if row["_adj"] >= 30 else ("0D47A1" if row["_adj"] >= 25 else fg))
        if is_closer or is_bustout: adj_color = fg

        row_data = [row["#"], row["Song"], row["Tier"], row["City Freq"],
                    row["Shows Since Last Played"], row["Adj Score"], row["Show Position"]]
        aligns = ["center", "left", "center", "center", "center", "center", "center"]

        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = bdr
            fc = adj_color if col == 6 else fg
            c.font = Font(name="Arial", size=10, color=fc, bold=(col in [2, 6]))
        ws.row_dimensions[r].height = 18

    lr = len(rows) + 6
    ws.merge_cells(f"A{lr}:G{lr}")
    c = ws.cell(row=lr, column=1,
        value="Green adj score ≥ 30%  |  Blue adj score ≥ 25%  |  Dark green = bustout (gap ≥ 150 shows)  |  Dark red = closer")
    c.font = Font(name="Arial", size=8, italic=True, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[lr].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Top 50 / Sphere helpers ─────────────────────────────────────────────────

API_KEY = "C5A172D7DD1198D7BB1C"

def _ssl_ctx():
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx

@st.cache_data(ttl=3600)
def fetch_sphere_songs_st():
    ctx = _ssl_ctx()
    sphere_dates = []
    with open(FILEPATH, newline="") as f:
        for row in csv.DictReader(f):
            if "sphere" in row.get("venue","").lower() and row["date"].startswith("2026"):
                sphere_dates.append(row["date"])
    current_year = datetime.date.today().year
    for year in [current_year, current_year-1]:
        url = f"https://api.phish.net/v5/shows/query.json?apikey={API_KEY}&year={year}"
        req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=10) as r:
                data = json.loads(r.read())
            for show in data.get("data",[]):
                if "sphere" in show.get("venue","").lower():
                    d = show.get("showdate","")
                    if d and d.startswith("2026") and d not in sphere_dates:
                        sphere_dates.append(d)
        except: pass

    sphere_dates = sorted(set(sphere_dates))
    song_dates = {}
    for date in sphere_dates:
        url = f"https://api.phish.net/v5/setlists/showdate/{date}.json?apikey={API_KEY}"
        req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=10) as r:
                data = json.loads(r.read())
            for s in data.get("data",[]):
                name = s["song"]
                song_dates.setdefault(name, [])
                if date not in song_dates[name]:
                    song_dates[name].append(date)
        except: pass
    return song_dates, sphere_dates

@st.cache_data(ttl=3600)
def build_top50_st():
    global_counter = Counter()
    global_shows = 0
    with open(FILEPATH, newline="") as f:
        for row in csv.DictReader(f):
            if not row["date"][:4].isdigit() or int(row["date"][:4]) < START_YEAR:
                continue
            songs = [s.strip() for s in row["setlist"].split("|") if s.strip()]
            if not songs: continue
            global_shows += 1
            global_counter.update(songs)
    return global_counter, global_shows

def _build_top50_xlsx_buf(top50, global_shows, sphere_songs):
    """Build the Top 50 xlsx in memory and return a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 50 Phish Songs"

    thin = Side(style="thin", color="333333")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    generated = datetime.datetime.now().strftime("%B %d, %Y  %I:%M %p")

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "PHISH  —  Top 50 Most Played Songs (2008–Present)"
    c.font      = Font(name="Arial", bold=True, size=16, color="F0E68C")
    c.fill      = PatternFill("solid", fgColor="1A1A2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value     = (f"Based on {global_shows} shows  |  % = chance on any given night  |  "
                   f"★ = played at Sphere  |  Updated: {generated}")
    c.font      = Font(name="Arial", italic=True, size=9, color="AAAACC")
    c.fill      = PatternFill("solid", fgColor="1A1A2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 5

    headers    = ["Rank", "Song", "Times Played", "% Any Night", "Sphere Dates", "YouTube"]
    col_widths = [7, 36, 16, 14, 28, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=11, color="F0E68C")
        c.fill      = PatternFill("solid", fgColor="1A1A2E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 22

    TIER_FILL_X = {
        "green":  ("1B4D1B", "90EE90"),
        "orange": ("4D2E00", "FFB347"),
        "purple": ("2E0050", "CE93D8"),
    }
    SPHERE_BG_X, SPHERE_FG_X = "7B5800", "FFFACD"

    for rank, (song, count) in enumerate(top50, 1):
        r   = rank + 4
        pct = count / global_shows * 100
        key = "green" if rank <= 10 else ("orange" if rank <= 25 else "purple")
        bg, fg = TIER_FILL_X[key]
        dates_played = sphere_songs.get(song, [])
        sphere_label = ("★  " + ",  ".join(d[5:] for d in sorted(dates_played))
                        if dates_played else "")
        yt_query = urllib.parse.quote_plus(f"Phish {song} Sphere Las Vegas 2026")
        yt_url   = f"https://www.youtube.com/results?search_query={yt_query}"

        row_data = [rank, song, count, f"{pct:.1f}%", sphere_label]
        aligns   = ["center", "left", "center", "center", "center"]
        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            c = ws.cell(row=r, column=col, value=val)
            if col == 5 and dates_played:
                c.fill = PatternFill("solid", fgColor=SPHERE_BG_X)
                c.font = Font(name="Arial", size=10, color=SPHERE_FG_X, bold=True)
            else:
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Arial", size=10, color=fg, bold=(col in [2, 4]))
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = bdr

        yt_cell = ws.cell(row=r, column=6)
        if dates_played:
            yt_cell.value     = "▶ Watch"
            yt_cell.hyperlink = yt_url
            yt_cell.fill      = PatternFill("solid", fgColor=SPHERE_BG_X)
            yt_cell.font      = Font(name="Arial", size=10, color="4FC3F7", bold=True, underline="single")
        else:
            yt_cell.fill = PatternFill("solid", fgColor=bg)
            yt_cell.font = Font(name="Arial", size=10, color=fg)
        yt_cell.alignment = Alignment(horizontal="center", vertical="center")
        yt_cell.border    = bdr
        ws.row_dimensions[r].height = 18

    lr = 57
    ws.merge_cells(f"A{lr}:F{lr}")
    c = ws.cell(row=lr, column=1,
                value="Green = Top 10  |  Orange = Ranks 11–25  |  Purple = Ranks 26–50  |  Gold = played at Sphere 2026")
    c.font      = Font(name="Arial", size=8, italic=True, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def ask_trey_st(question, global_counter, global_shows):
    all_songs = list(global_counter.keys())
    q_lower = question.lower()
    match = None
    best_score = 0
    for song in all_songs:
        if song.lower() in q_lower and len(song) > best_score:
            match = song
            best_score = len(song)
    if not match:
        words = question.replace("?","").replace(",","").split()
        for length in range(min(6,len(words)),0,-1):
            for start in range(len(words)-length+1):
                phrase = " ".join(words[start:start+length])
                close = difflib.get_close_matches(phrase, all_songs, n=1, cutoff=0.55)
                if close:
                    match = close[0]; break
            if match: break
    if not match:
        return None, "I didn't catch a song name — try something like 'Will you play Sand?' and I'll give you the real numbers.", {}

    count = global_counter[match]
    pct   = count / global_shows * 100
    sphere_songs, sphere_dates = fetch_sphere_songs_st()
    sphere_played = sphere_songs.get(match, [])
    shows_done  = len([d for d in sphere_dates if d <= datetime.date.today().isoformat()])
    shows_left  = len(sphere_dates) - shows_done

    song_last_idx = {}
    show_idx = 0
    with open(FILEPATH, newline="") as f:
        for row in csv.DictReader(f):
            if not row["date"][:4].isdigit() or int(row["date"][:4]) < START_YEAR:
                continue
            songs_row = [s.strip() for s in row["setlist"].split("|") if s.strip()]
            if not songs_row:
                continue  # skip rows with no setlist (future/scheduled shows)
            for s in songs_row:
                song_last_idx[s] = show_idx
            show_idx += 1
    gap = max(0, global_shows - song_last_idx.get(match, 0) - 1)
    log_denom = math.log(global_shows + 1) if global_shows > 0 else 1
    adj_pct = min(pct * (1 + math.log(gap + 1) / log_denom), 99.9)

    if sphere_played:
        if len(sphere_played) > 1:
            answer = f"Ha — {match}! We've already played that {len(sphere_played)} times this run ({', '.join(d[5:] for d in sphere_played)}). Historically it shows up in {pct:.1f}% of our shows."
        else:
            answer = f"{match} — yeah, we played that on {sphere_played[0][5:]}. It's in our setlist about {pct:.1f}% of the time overall. Could we revisit it? Maybe — but we don't like to repeat too much in a single run."
    elif gap > 200:
        answer = f"{match} — oh man. That one's been sitting in the vault for {gap} shows. Historically {pct:.1f}% of nights, but gap-adjusted it's up around {adj_pct:.1f}%. We've got {shows_left} nights left at the Sphere... keep your ears open."
    elif pct >= 15:
        answer = f"{match} is basically a staple — we play it {pct:.1f}% of shows. It hasn't shown up yet this Sphere run and we've got {shows_left} nights left. I'd feel confident betting on this one."
    elif pct >= 5:
        answer = f"Good question. {match} shows up in about {pct:.1f}% of our shows — a song we genuinely love. Last played {gap} shows ago. Adjusted probability: {adj_pct:.1f}%. There's a real shot."
    elif pct >= 1:
        answer = f"{match} — that's a deep cut. About {pct:.1f}% of shows, {gap} shows since we last played it. These Sphere shows feel special though. Adjusted probability: {adj_pct:.1f}%. Stranger things have happened."
    else:
        answer = f"Oh wow, {match} — now THAT would be something. We've played it {count} time{'s' if count!=1 else ''} since 2008. The Sphere feels like the right place to dust off something unexpected."

    is_bust_out = gap > 500 and not sphere_played
    if is_bust_out:
        answer = f"⭐ **BUST OUT** — {answer}"
    stats = {"pct": round(pct,1), "gap": gap, "adj": round(adj_pct,1),
             "sphere": [d[5:] for d in sphere_played], "bust_out": is_bust_out}
    return match, answer, stats


# ── UI ──────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Gotta-Jibbootistics", page_icon="🍩", layout="centered")

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&family=Shrikhand&family=Kaushan+Script&family=JetBrains+Mono:wght@400;600&display=swap');

    /* Softer, brighter base with a more visible donut pattern */
    .stApp {
        background-color: #24243e;
        background-image:
            radial-gradient(circle at center, transparent 24%, rgba(255, 201, 150, 0.12) 24%, rgba(255, 201, 150, 0.12) 46%, transparent 46%),
            radial-gradient(circle at center, transparent 24%, rgba(140, 180, 220, 0.10) 24%, rgba(140, 180, 220, 0.10) 46%, transparent 46%),
            linear-gradient(135deg, #2a2a48 0%, #1e1e36 50%, #2a2a48 100%);
        background-size: 160px 160px, 160px 160px, 100% 100%;
        background-position: 0 0, 80px 80px, 0 0;
        background-attachment: fixed;
    }

    /* Hide Streamlit's default top toolbar so it doesn't cover the hero */
    header[data-testid="stHeader"] {
        background: transparent !important;
        height: 0 !important;
    }
    #MainMenu { visibility: hidden; }

    /* Dark-theme expanders so they don't show as white bars on the dark page */
    div[data-testid="stExpander"],
    div[data-testid="stExpander"] > details,
    div[data-testid="stExpander"] details > summary {
        background: rgba(20, 20, 36, 0.55) !important;
        border: 1px solid rgba(255, 255, 255, 0.06) !important;
        border-radius: 10px !important;
        color: #F5F5F7 !important;
    }
    div[data-testid="stExpander"] details > summary {
        border: none !important;
        color: #FFE98A !important;
    }
    div[data-testid="stExpander"] details > summary:hover {
        color: #FFF3B0 !important;
    }
    div[data-testid="stExpander"] details[open] > summary {
        border-bottom: 1px solid rgba(255, 255, 255, 0.06) !important;
        border-radius: 10px 10px 0 0 !important;
    }
    div[data-testid="stExpander"] svg { fill: #FFE98A !important; }

    /* Refined content card */
    .block-container {
        background: rgba(20, 20, 36, 0.78);
        border-radius: 16px;
        padding: 3.5rem 2.5rem 2.5rem 2.5rem !important;
        margin-top: 1.5rem !important;
        border: 1px solid rgba(255, 255, 255, 0.06);
        box-shadow: 0 8px 40px rgba(0, 0, 0, 0.4);
        backdrop-filter: blur(8px);
    }

    /* Typography — Outfit for a modern, refined feel */
    html, body, [class*="stApp"], .stMarkdown, .stMarkdown p,
    .stMarkdown li, .stMarkdown span, label, .stCaption,
    div[data-testid="stMarkdownContainer"] p {
        color: #F5F5F7 !important;
        font-family: 'Outfit', -apple-system, "Helvetica Neue", Arial, sans-serif !important;
        font-weight: 400 !important;
        line-height: 1.55;
    }
    h1 {
        color: #FFF3B0 !important;
        font-family: 'Shrikhand', 'Kaushan Script', cursive !important;
        font-weight: 400 !important;
        letter-spacing: 0.01em;
        font-size: 2.8rem !important;
        margin-bottom: 0.3rem !important;
    }
    h2, h3, h4 {
        color: #FFE98A !important;
        font-family: 'Shrikhand', 'Kaushan Script', cursive !important;
        font-weight: 400 !important;
        letter-spacing: 0.01em;
    }
    h2 { font-size: 2rem !important; }
    h3 { font-size: 1.5rem !important; }
    h4 { font-size: 1.2rem !important; }
    code, pre { font-family: 'JetBrains Mono', monospace !important; }
    .stMarkdown small, .stCaption, div[data-testid="stCaptionContainer"] {
        color: #B8B8C8 !important;
    }

    /* Inputs */
    .stTextInput > div > div > input,
    .stSelectbox > div > div {
        background-color: #1e1e34 !important;
        color: #FFF3B0 !important;
        border: 1px solid rgba(255, 255, 255, 0.12) !important;
        border-radius: 8px !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: #FFD166 !important;
        box-shadow: 0 0 0 2px rgba(255, 209, 102, 0.2) !important;
    }

    /* Buttons */
    .stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {
        background: linear-gradient(135deg, #2a2a4a 0%, #1e1e34 100%) !important;
        color: #FFF3B0 !important;
        border: 1px solid rgba(255, 209, 102, 0.3) !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        transition: all 0.15s ease !important;
    }
    .stButton > button:hover, .stDownloadButton > button:hover, .stFormSubmitButton > button:hover {
        border-color: #FFD166 !important;
        box-shadow: 0 0 12px rgba(255, 209, 102, 0.25) !important;
        transform: translateY(-1px);
    }

    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        border-bottom: 1px solid rgba(255, 255, 255, 0.08);
    }
    .stTabs [data-baseweb="tab"] {
        color: #B8B8C8 !important;
        font-weight: 400 !important;
        padding: 10px 16px !important;
        font-family: 'Shrikhand', 'Kaushan Script', cursive !important;
        font-size: 1rem !important;
        letter-spacing: 0.02em;
    }
    .stTabs [aria-selected="true"] {
        color: #FFF3B0 !important;
    }

    /* Tables & dividers */
    table { border-radius: 10px; overflow: hidden; }
    hr { border-color: rgba(255, 255, 255, 0.08) !important; }

    /* Alerts softened */
    div[data-testid="stAlert"] {
        border-radius: 10px;
        background: rgba(30, 30, 52, 0.85) !important;
    }

    /* ── Card & metric system ──────────────────────── */
    .gj-card {
        background: linear-gradient(135deg, rgba(36, 36, 62, 0.92) 0%, rgba(30, 30, 52, 0.92) 100%);
        border: 1px solid rgba(255, 243, 176, 0.10);
        border-radius: 14px;
        padding: 20px 24px;
        margin: 16px 0;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.35);
    }
    .gj-card-accent {
        border-left: 3px solid #FFD166;
    }

    /* Hero */
    .gj-hero {
        display: flex; align-items: center; gap: 18px;
        margin: -8px 0 4px 0;
    }
    .gj-hero-logo { flex: 0 0 auto; }
    .gj-hero-text { flex: 1 1 auto; }
    .gj-hero h1 {
        margin: 0 0 2px 0 !important;
        font-size: 2.6rem !important;
        line-height: 1.1;
        font-family: 'Shrikhand', 'Kaushan Script', cursive !important;
    }
    .gj-hero-tag {
        color: #F4A88E;
        font-style: italic;
        font-size: 13px;
        letter-spacing: 0.01em;
        margin: 0;
    }

    /* Metrics bar */
    .gj-metrics {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 12px;
        margin: 14px 0 20px 0;
    }
    .gj-metric {
        background: linear-gradient(135deg, rgba(46, 46, 76, 0.85) 0%, rgba(30, 30, 52, 0.85) 100%);
        border: 1px solid rgba(255, 243, 176, 0.08);
        border-radius: 10px;
        padding: 12px 14px;
        text-align: left;
    }
    .gj-metric-label {
        color: #B8B8C8;
        font-size: 11px;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 500;
        margin-bottom: 4px;
    }
    .gj-metric-value {
        color: #FFF3B0;
        font-size: 1.6rem;
        font-weight: 700;
        line-height: 1.1;
        font-family: 'Outfit', sans-serif;
    }
    .gj-metric-sub {
        color: #8888a0;
        font-size: 10.5px;
        margin-top: 3px;
    }
    @media (max-width: 720px) {
        .gj-metrics { grid-template-columns: repeat(2, 1fr); }
    }

    /* Section headings inside cards */
    .gj-section-head {
        color: #FFE98A;
        font-family: 'Shrikhand', 'Kaushan Script', cursive;
        font-weight: 400;
        font-size: 1.5rem;
        margin: 0 0 6px 0;
        letter-spacing: 0.01em;
    }
    .gj-section-sub {
        color: #9a9ab0;
        font-size: 12px;
        margin: 0 0 14px 0;
    }

    /* Methodology footer */
    .gj-footer {
        margin-top: 32px;
        padding-top: 18px;
        border-top: 1px solid rgba(255, 255, 255, 0.08);
        color: #8888a0;
        font-size: 12px;
        text-align: center;
    }

    /* ── Animations ──────────────────────────── */
    @property --count {
        syntax: '<integer>';
        initial-value: 0;
        inherits: false;
    }

    /* Staggered fade-slide for metrics */
    .gj-metric {
        opacity: 0;
        transform: translateY(8px);
        animation: gj-enter 0.55s cubic-bezier(0.2, 0.7, 0.3, 1) forwards;
    }
    .gj-metric:nth-child(1) { animation-delay: 0.05s; }
    .gj-metric:nth-child(2) { animation-delay: 0.15s; }
    .gj-metric:nth-child(3) { animation-delay: 0.25s; }
    .gj-metric:nth-child(4) { animation-delay: 0.35s; }
    @keyframes gj-enter {
        to { opacity: 1; transform: translateY(0); }
    }

    /* Hero logo gentle spin-in */
    .gj-hero-logo svg {
        animation: gj-logo-in 0.8s cubic-bezier(0.2, 0.7, 0.3, 1);
    }
    @keyframes gj-logo-in {
        from { opacity: 0; transform: rotate(-25deg) scale(0.7); }
        to   { opacity: 1; transform: rotate(0) scale(1); }
    }

    /* Share link popup */
    .gj-share {
        background: rgba(40, 40, 70, 0.85);
        border: 1px solid rgba(255, 209, 102, 0.3);
        border-radius: 8px;
        padding: 10px 14px;
        font-family: 'JetBrains Mono', monospace;
        font-size: 12px;
        color: #FFF3B0;
        word-break: break-all;
        margin: 10px 0;
    }

    /* Reduced motion respect */
    @media (prefers-reduced-motion: reduce) {
        .gj-metric, .gj-hero-logo svg { animation: none !important; opacity: 1 !important; transform: none !important; }
        .gj-count::before { animation: none !important; }
    }
    </style>
""", unsafe_allow_html=True)


# ── Logo + Hero helpers ─────────────────────────────────────────

DONUT_LOGO_SVG = (
    '<svg width="64" height="64" viewBox="0 0 64 64" xmlns="http://www.w3.org/2000/svg" style="display:block">'
    '<defs>'
    '<radialGradient id="dough" cx="40%" cy="35%" r="65%">'
    '<stop offset="0%" stop-color="#f4b8a0"/>'
    '<stop offset="100%" stop-color="#d87a62"/>'
    '</radialGradient>'
    '<radialGradient id="frost" cx="45%" cy="35%" r="70%">'
    '<stop offset="0%" stop-color="#FFE098"/>'
    '<stop offset="100%" stop-color="#E8806E"/>'
    '</radialGradient>'
    '</defs>'
    '<circle cx="32" cy="32" r="26" fill="url(#dough)" stroke="#a85a48" stroke-width="1"/>'
    '<path d="M 32 8 A 24 24 0 0 1 32 56 A 24 24 0 0 1 32 8 Z M 32 14 A 18 18 0 0 0 32 50 A 18 18 0 0 0 32 14 Z" fill="url(#frost)" opacity="0.95"/>'
    '<circle cx="32" cy="32" r="9" fill="#24243e"/>'
    '<rect x="18" y="18" width="2.5" height="7" rx="1.2" fill="#FFF3B0" transform="rotate(25 19.25 21.5)"/>'
    '<rect x="44" y="21" width="2.5" height="7" rx="1.2" fill="#8fd8f0" transform="rotate(-30 45.25 24.5)"/>'
    '<rect x="21" y="42" width="2.5" height="7" rx="1.2" fill="#d4a8e0" transform="rotate(55 22.25 45.5)"/>'
    '<rect x="41" y="41" width="2.5" height="7" rx="1.2" fill="#FFD166" transform="rotate(-50 42.25 44.5)"/>'
    '<rect x="15" y="32" width="2.5" height="7" rx="1.2" fill="#B9F6CA" transform="rotate(90 16.25 35.5)"/>'
    '</svg>'
)


@st.cache_data(ttl=600)
def get_hero_stats():
    """Compute headline numbers for the hero metrics row."""
    global_counter, global_shows, _, _, total_shows, _ = load_data()
    # Sphere progress
    try:
        sphere_songs, sphere_dates = fetch_sphere_songs_st()
    except Exception:
        sphere_songs, sphere_dates = {}, []
    today_iso = datetime.date.today().isoformat()
    done = sum(1 for d in sphere_dates if d <= today_iso)
    left = max(0, len(sphere_dates) - done)
    return {
        "shows": global_shows,
        "songs": len(global_counter),
        "sphere_done": done,
        "sphere_left": left,
        "unique_sphere_songs": len(sphere_songs),
    }


def _counter_css(idx, target):
    """Legacy — kept for backwards compat, now unused."""
    return ""


def render_hero():
    """Render SVG logo + wordmark + tagline + metrics row + flying hotdogs."""
    stats = get_hero_stats()
    remaining_plural = 's' if stats['sphere_left'] != 1 else ''
    total_sphere = stats['sphere_done'] + stats['sphere_left']

    # Flying hotdogs (MSG NYE nod). 8 dogs, varied speeds/arcs, full-screen traversal.
    hotdog_css = """
    <style>
    .gj-hotdogs { position: fixed; inset: 0; pointer-events: none; z-index: 9999;
                  overflow: hidden; }
    .gj-hotdog { position: absolute; opacity: 0; will-change: transform; }
    @media (prefers-reduced-motion: reduce) { .gj-hotdog { display: none; } }

    /* Each dog has its own arc, speed, and spin. Start at -15vw, end past 120vw. */
    @keyframes gj-fly-1 { /* fast low arc */
        0%   { opacity: 0; transform: translate(-15vw, 0)    rotate(0deg)    scale(0.9); }
        8%   { opacity: 1; }
        50%  { transform: translate(50vw, -8vh) rotate(360deg) scale(1.05); }
        92%  { opacity: 1; }
        100% { opacity: 0; transform: translate(125vw, -4vh)  rotate(720deg)  scale(0.9); }
    }
    @keyframes gj-fly-2 { /* slow high arc */
        0%   { opacity: 0; transform: translate(-15vw, 0)    rotate(0deg)    scale(0.85); }
        10%  { opacity: 1; }
        50%  { transform: translate(50vw, -22vh) rotate(540deg) scale(1.15); }
        90%  { opacity: 1; }
        100% { opacity: 0; transform: translate(125vw, -40vh) rotate(1080deg) scale(0.8); }
    }
    @keyframes gj-fly-3 { /* medium dipping */
        0%   { opacity: 0; transform: translate(-15vw, 0)    rotate(0deg)    scale(1); }
        10%  { opacity: 1; }
        50%  { transform: translate(50vw, 12vh)  rotate(-360deg) scale(1.1); }
        90%  { opacity: 1; }
        100% { opacity: 0; transform: translate(125vw, 0)     rotate(-720deg) scale(1); }
    }
    @keyframes gj-fly-4 { /* fast straight */
        0%   { opacity: 0; transform: translate(-15vw, 0)    rotate(0deg)    scale(0.95); }
        8%   { opacity: 1; }
        50%  { transform: translate(55vw, -2vh) rotate(180deg) scale(1); }
        92%  { opacity: 1; }
        100% { opacity: 0; transform: translate(125vw, 4vh)   rotate(360deg)  scale(0.95); }
    }
    @keyframes gj-fly-5 { /* slow big arc up */
        0%   { opacity: 0; transform: translate(-15vw, 0)    rotate(0deg)    scale(0.9); }
        12%  { opacity: 1; }
        50%  { transform: translate(50vw, -30vh) rotate(900deg) scale(1.25); }
        88%  { opacity: 1; }
        100% { opacity: 0; transform: translate(125vw, -50vh) rotate(1440deg) scale(0.75); }
    }
    @keyframes gj-fly-6 { /* zig-zag */
        0%   { opacity: 0; transform: translate(-15vw, 0)    rotate(0deg)    scale(1); }
        10%  { opacity: 1; }
        30%  { transform: translate(25vw, -10vh) rotate(180deg); }
        60%  { transform: translate(60vw, 8vh)   rotate(-180deg); }
        90%  { opacity: 1; }
        100% { opacity: 0; transform: translate(125vw, -5vh) rotate(540deg) scale(1); }
    }

    .gj-hd1 { top: 10%; left: 0; font-size: 2.2rem;
              animation: gj-fly-1 2.6s cubic-bezier(.3,.6,.4,1) 0.05s forwards; }
    .gj-hd2 { top: 22%; left: 0; font-size: 3.2rem;
              animation: gj-fly-2 5.2s cubic-bezier(.25,.5,.5,1) 0.4s forwards; }
    .gj-hd3 { top: 36%; left: 0; font-size: 2.6rem;
              animation: gj-fly-3 3.8s cubic-bezier(.3,.6,.4,1) 0.15s forwards; }
    .gj-hd4 { top: 50%; left: 0; font-size: 2rem;
              animation: gj-fly-4 2.2s cubic-bezier(.35,.7,.4,1) 0.6s forwards; }
    .gj-hd5 { top: 64%; left: 0; font-size: 2.8rem;
              animation: gj-fly-5 4.6s cubic-bezier(.2,.5,.5,1) 0.25s forwards; }
    .gj-hd6 { top: 76%; left: 0; font-size: 2.4rem;
              animation: gj-fly-6 3.4s cubic-bezier(.3,.6,.4,1) 0.75s forwards; }
    .gj-hd7 { top: 86%; left: 0; font-size: 3rem;
              animation: gj-fly-2 5s   cubic-bezier(.25,.5,.5,1) 0.1s forwards; }
    .gj-hd8 { top: 94%; left: 0; font-size: 2rem;
              animation: gj-fly-1 2.4s cubic-bezier(.35,.7,.4,1) 0.9s forwards; }
    </style>
    <div class="gj-hotdogs">
      <span class="gj-hotdog gj-hd1">🌭</span>
      <span class="gj-hotdog gj-hd2">🌭</span>
      <span class="gj-hotdog gj-hd3">🌭</span>
      <span class="gj-hotdog gj-hd4">🌭</span>
      <span class="gj-hotdog gj-hd5">🌭</span>
      <span class="gj-hotdog gj-hd6">🌭</span>
      <span class="gj-hotdog gj-hd7">🌭</span>
      <span class="gj-hotdog gj-hd8">🌭</span>
    </div>
    """
    st.markdown(hotdog_css, unsafe_allow_html=True)

    # Metric values: rendered directly with a fade-up animation — reliable everywhere.
    metric_anim_css = """
    <style>
    .gj-metric-value { animation: gj-metric-in 0.8s cubic-bezier(0.2,0.7,0.3,1) 0.15s both; }
    .gj-metric:nth-child(2) .gj-metric-value { animation-delay: 0.25s; }
    .gj-metric:nth-child(3) .gj-metric-value { animation-delay: 0.35s; }
    .gj-metric:nth-child(4) .gj-metric-value { animation-delay: 0.45s; }
    @keyframes gj-metric-in {
        from { opacity: 0; transform: translateY(8px) scale(0.96); }
        to   { opacity: 1; transform: translateY(0)   scale(1); }
    }
    </style>
    """

    # Build all on one line per logical block; avoids markdown 4-space code-block trap.
    html_parts = [
        metric_anim_css,
        '<div class="gj-hero">',
        f'<div class="gj-hero-logo">{DONUT_LOGO_SVG}</div>',
        '<div class="gj-hero-text">',
        '<h1>Gotta-Jibbootistics</h1>',
        '<p class="gj-hero-tag">whatever you do, take care of your shoes</p>',
        '</div></div>',
        '<div class="gj-metrics">',
        f'<div class="gj-metric"><div class="gj-metric-label">Shows Analyzed</div>'
        f'<div class="gj-metric-value">{stats["shows"]:,}</div>'
        f'<div class="gj-metric-sub">2008 – present</div></div>',
        f'<div class="gj-metric"><div class="gj-metric-label">Unique Songs</div>'
        f'<div class="gj-metric-value">{stats["songs"]:,}</div>'
        f'<div class="gj-metric-sub">in the catalog</div></div>',
        f'<div class="gj-metric"><div class="gj-metric-label">Sphere 2026</div>'
        f'<div class="gj-metric-value">{stats["sphere_done"]} '
        f'<span style="color:#8888a0;font-weight:400;font-size:1rem">/ {total_sphere}</span></div>'
        f'<div class="gj-metric-sub">{stats["sphere_left"]} show{remaining_plural} remaining</div></div>',
        f'<div class="gj-metric"><div class="gj-metric-label">Sphere Setlist</div>'
        f'<div class="gj-metric-value">{stats["unique_sphere_songs"]:,}</div>'
        f'<div class="gj-metric-sub">unique songs so far</div></div>',
        '</div>',
    ]
    # Use st.html() (Streamlit 1.33+) — renders raw HTML without markdown parsing.
    # This avoids any 4-space / HTML-block-boundary edge cases in the markdown parser.
    try:
        st.html("".join(html_parts))
    except AttributeError:
        # Fallback for older Streamlit versions
        st.markdown("".join(html_parts), unsafe_allow_html=True)


# ── Share-link helpers ─────────────────────────────────────
def _build_share_url(params: dict) -> str:
    """Build a shareable URL with current query params."""
    base = "https://sphereprob.streamlit.app"
    qs = urllib.parse.urlencode({k: v for k, v in params.items() if v})
    return f"{base}/?{qs}" if qs else base


def render_share_box(params: dict, key: str):
    """Inline share-link display — shows a monospace URL on click."""
    col_a, col_b = st.columns([1, 3])
    with col_a:
        show = st.button("🔗 Share this prediction", key=f"share_{key}")
    if show:
        st.session_state[f"share_open_{key}"] = True
    if st.session_state.get(f"share_open_{key}"):
        url = _build_share_url(params)
        st.markdown(
            f'<div class="gj-share">{url}</div>'
            f'<div style="font-size:11px;color:#8888a0;margin-top:-4px">'
            f'Copy this URL — anyone with the link will see the same prediction.</div>',
            unsafe_allow_html=True,
        )


def render_methodology_footer():
    """Collapsible methodology + about section."""
    with st.expander("About the methodology", expanded=False):
        st.markdown("""
        **How predictions are scored**

        - **City Frequency** — how often a song has been played in the target city (or Las Vegas for Sphere predictions).
        - **Gap Boost** — songs that haven't been played in many shows get a multiplier: `1 + log(gap+1) / log(total_shows+1)`. Longer droughts = higher chance of a bustout.
        - **Tier System** — songs are classified by global frequency: Staple (>15%), Common (5–15%), Occasional (1–5%), Rare (<1%). Each predicted setlist fills a realistic mix of tiers.
        - **Position Ordering** — each song has a typical place in the set (opener, mid-show, closer), derived from its average position across all shows. One opener + one closer is enforced per predicted setlist.
        - **Recent Rotation Boost** *(Sphere tab only)* — songs the band has played in the 10–15 shows immediately prior to the target date get a 1.6× multiplier, capturing songs currently "in rotation."

        **Sphere-specific logic**

        - Songs already played during the 2026 Sphere run are excluded from future-show predictions — the band rarely repeats within a single run.
        - Sphere setlists are fetched live from the phish.net API and cached for an hour.

        **Data source**

        Setlists scraped from [phish.net](https://phish.net) covering 2008 through the current tour. Updated weekly.
        """)
    st.markdown(
        '<div class="gj-footer">Gotta-Jibbootistics · by fans for fans · '
        'data via <a href="https://phish.net" style="color:#8fd8f0">phish.net</a></div>',
        unsafe_allow_html=True
    )

render_hero()


@st.cache_data(ttl=1800)
def _compute_last_night_accuracy():
    """Compute accuracy metrics for the most recently played Sphere show.
    Returns dict or None if no prior show / not enough data."""
    try:
        sphere_songs, sphere_dates = fetch_sphere_songs_st()
    except Exception:
        return None
    today_iso = datetime.date.today().isoformat()
    played = sorted(d for d in sphere_dates if d <= today_iso)
    if not played:
        return None
    # Walk back from today: skip any "scheduled but not yet played" dates
    # (the API lists them but they have no setlist data yet).
    last_show, actual = None, []
    for d in reversed(played):
        songs_for_d = sorted([s for s, ds in sphere_songs.items() if d in ds])
        if songs_for_d:
            last_show, actual = d, songs_for_d
            break
    if not last_show:
        return None
    prior = {s: [d for d in ds if d < last_show] for s, ds in sphere_songs.items()}
    prior = {s: ds for s, ds in prior.items() if ds}
    try:
        retro = generate_sphere_setlist(last_show, prior)
        predicted = [r["Song"] for r in retro["rows"]]
    except Exception:
        return None
    if not predicted:
        return None
    pred_set = set(predicted)
    act_set  = set(actual)
    hits     = pred_set & act_set
    return {
        "date": last_show,
        "predicted": predicted,
        "actual": actual,
        "hits": sorted(hits),
        "misses": sorted(pred_set - act_set),
        "precision": len(hits) / len(pred_set),
        "recall": len(hits) / len(act_set),
    }


def render_last_night_accuracy_banner():
    """Prominent banner on the main page showing last night's prediction accuracy."""
    acc = _compute_last_night_accuracy()
    if not acc:
        return

    pretty = datetime.date.fromisoformat(acc["date"]).strftime("%b %d")
    prec = acc["precision"] * 100
    rec  = acc["recall"] * 100
    # Color tone based on precision
    if prec >= 50:
        tone, accent = "#81C784", "rgba(129, 199, 132, 0.5)"
    elif prec >= 25:
        tone, accent = "#FFD54F", "rgba(255, 213, 79, 0.5)"
    else:
        tone, accent = "#FF8A65", "rgba(255, 138, 101, 0.5)"

    hit_pills = "".join(
        f'<span style="background:rgba(129,199,132,0.15);color:#B9F6CA;'
        f'padding:3px 9px;border-radius:12px;font-size:11.5px;margin:2px 3px;'
        f'display:inline-block;border:1px solid rgba(129,199,132,0.3)">✓ {s}</span>'
        for s in acc["hits"]
    ) or '<span style="color:#666;font-size:11px">None this time</span>'

    miss_pills = "".join(
        f'<span style="background:rgba(255,138,101,0.12);color:#F4A88E;'
        f'padding:3px 9px;border-radius:12px;font-size:11.5px;margin:2px 3px;'
        f'display:inline-block;border:1px solid rgba(255,138,101,0.25)">✗ {s}</span>'
        for s in acc["misses"]
    ) or '<span style="color:#666;font-size:11px">No misses</span>'

    st.markdown(f"""
    <style>
    .gj-accuracy-banner {{
        background: linear-gradient(135deg, rgba(46,46,76,0.92) 0%, rgba(30,30,52,0.92) 100%);
        border: 1px solid rgba(255,255,255,0.08);
        border-left: 4px solid {tone};
        border-radius: 14px;
        padding: 18px 22px;
        margin: 18px 0 22px 0;
        box-shadow: 0 4px 24px rgba(0,0,0,0.35), 0 0 0 1px {accent};
        animation: gj-acc-pulse 0.9s ease-out;
    }}
    @keyframes gj-acc-pulse {{
        from {{ transform: translateY(6px); opacity: 0; }}
        to   {{ transform: translateY(0);   opacity: 1; }}
    }}
    .gj-acc-title {{
        color: {tone};
        font-size: 12.5px;
        font-weight: 700;
        letter-spacing: 0.12em;
        text-transform: uppercase;
        margin-bottom: 10px;
    }}
    .gj-acc-grid {{
        display: grid;
        grid-template-columns: 1.2fr 1fr 1fr;
        gap: 18px;
        align-items: center;
    }}
    .gj-acc-big {{
        color: #FFF3B0;
        font-family: 'Shrikhand', cursive;
        font-size: 2.6rem;
        line-height: 1;
        margin-bottom: 2px;
    }}
    .gj-acc-label {{
        color: #9a9ab0;
        font-size: 10.5px;
        text-transform: uppercase;
        letter-spacing: 0.08em;
    }}
    .gj-acc-sub {{
        color: #c8c8dc;
        font-size: 11.5px;
        margin-top: 2px;
    }}
    </style>
    <div class="gj-accuracy-banner">
        <div class="gj-acc-title">📊 Last Night's Prediction — {pretty}</div>
        <div class="gj-acc-grid">
            <div>
                <div class="gj-acc-big">{len(acc['hits'])}<span style="color:#666;font-size:1.2rem"> / {len(acc['predicted'])}</span></div>
                <div class="gj-acc-label">Hits</div>
                <div class="gj-acc-sub">{len(acc['actual'])} songs played · {len(acc['predicted'])} predicted</div>
            </div>
            <div>
                <div class="gj-acc-big" style="color:{tone}">{prec:.0f}%</div>
                <div class="gj-acc-label">Precision</div>
                <div class="gj-acc-sub">predicted songs that played</div>
            </div>
            <div>
                <div class="gj-acc-big" style="color:{tone}">{rec:.0f}%</div>
                <div class="gj-acc-label">Recall</div>
                <div class="gj-acc-sub">of setlist we predicted</div>
            </div>
        </div>
        <details style="margin-top:14px">
            <summary style="color:#c8c8dc;font-size:12px;cursor:pointer;user-select:none;
                            padding:6px 0;font-weight:500">
                See hits &amp; misses
            </summary>
            <div style="margin-top:8px">
                <div style="color:#B9F6CA;font-size:11px;font-weight:700;letter-spacing:0.06em;margin-bottom:5px">✓ HITS ({len(acc['hits'])})</div>
                <div>{hit_pills}</div>
                <div style="color:#F4A88E;font-size:11px;font-weight:700;letter-spacing:0.06em;margin:12px 0 5px 0">✗ MISSES ({len(acc['misses'])})</div>
                <div>{miss_pills}</div>
            </div>
        </details>
    </div>
    """, unsafe_allow_html=True)


render_last_night_accuracy_banner()


# ─────────────────────────────────────────────────────────────
# Lenny the Lizard — Clippy-style mascot in bottom-right corner
# Offers rotating tips. Click → jump+spin + new tip.
# ─────────────────────────────────────────────────────────────
def render_lenny_lizard():
    import streamlit.components.v1 as _components
    _components.html(r"""
    <script>
    (function() {
        const doc = window.parent.document;
        if (doc.__lennyInit) return;
        doc.__lennyInit = true;

        const TIPS = [
            "Welcome to Gotta-Jibbootistics! I'm Lenny. Click me for tips!",
            "🎸 The City Predictor builds a setlist for any city Phish has played.",
            "🏟️ Check the Top 50 tab to see Phish's most-played songs since 2008.",
            "🔮 The Sphere Predictor uses Las Vegas-specific data + recent rotation.",
            "🎲 Try the 3×3 Quick Bingo for a fast game — perfect for one-set bets!",
            "★ A FREE square sits in the center of every bingo card.",
            "🚨 Songs gone 500+ shows get flagged as BUST OUTs (max 1 per setlist).",
            "📊 Last night's prediction accuracy is right under the title.",
            "🍩 Curious about a song? Ask Trey on the Top 50 tab!",
            "⬇️ All bingo cards download as printable PDFs.",
            "🟡 Yellow = top picks · 🔵 Blue = common · 🟣 Purple = rare deep cuts.",
            "Did you know? You Enjoy Myself is the most-played Phish song.",
            "Tap a tab and watch the glowsticks rain. 🌟",
        ];

        // Build mascot
        const wrap = doc.createElement('div');
        wrap.id = 'lenny-wrap';
        wrap.innerHTML = `
            <div id="lenny-bubble">
                <span id="lenny-text"></span>
                <button id="lenny-close" aria-label="Close">×</button>
            </div>
            <div id="lenny-char" title="Click me for a tip!">
              <svg id="lenny-svg" viewBox="0 0 260 180" width="106" height="73" xmlns="http://www.w3.org/2000/svg">
                <defs>
                  <radialGradient id="lz-body" cx="50%" cy="20%" r="85%">
                    <stop offset="0%"  stop-color="#dff7a4"/>
                    <stop offset="40%" stop-color="#86c84e"/>
                    <stop offset="80%" stop-color="#3a8a25"/>
                    <stop offset="100%" stop-color="#0e3a0e"/>
                  </radialGradient>
                  <radialGradient id="lz-head" cx="40%" cy="25%" r="80%">
                    <stop offset="0%"  stop-color="#e4faae"/>
                    <stop offset="55%" stop-color="#82c64a"/>
                    <stop offset="100%" stop-color="#1f5a18"/>
                  </radialGradient>
                  <radialGradient id="lz-belly" cx="50%" cy="35%" r="65%">
                    <stop offset="0%"  stop-color="#fffbe6"/>
                    <stop offset="60%" stop-color="#fff099"/>
                    <stop offset="100%" stop-color="#c79a14"/>
                  </radialGradient>
                  <radialGradient id="lz-eye" cx="38%" cy="28%" r="85%">
                    <stop offset="0%"  stop-color="#fff8c2"/>
                    <stop offset="55%" stop-color="#ffcc33"/>
                    <stop offset="100%" stop-color="#a76d00"/>
                  </radialGradient>
                  <radialGradient id="lz-leg" cx="50%" cy="18%" r="85%">
                    <stop offset="0%"  stop-color="#b9e36e"/>
                    <stop offset="100%" stop-color="#1f5a18"/>
                  </radialGradient>
                  <linearGradient id="lz-tongue" x1="0" x2="1">
                    <stop offset="0%"  stop-color="#ff5fa2"/>
                    <stop offset="100%" stop-color="#c11860"/>
                  </linearGradient>
                </defs>

                <!-- ── TAIL: long curling shape on the right (animated wag) ── -->
                <g id="lz-tail-grp">
                  <path d="M 175,118
                           Q 230,118 240,80
                           Q 245,40 215,42
                           Q 235,55 220,80
                           Q 205,100 175,102 Z"
                        fill="url(#lz-body)" stroke="#0e3a0e" stroke-width="1"/>
                  <!-- yellow dots running down the tail -->
                  <circle cx="200" cy="100" r="2.6" fill="#ffd54f"/>
                  <circle cx="215" cy="92"  r="2.6" fill="#ffd54f"/>
                  <circle cx="226" cy="78"  r="2.4" fill="#ffd54f"/>
                  <circle cx="226" cy="62"  r="2.2" fill="#ffd54f"/>
                  <circle cx="218" cy="50"  r="2"   fill="#ffd54f"/>
                </g>

                <!-- ── BACK LEG (right side, bent outward) ── -->
                <g transform="rotate(-12 165 138)">
                  <ellipse cx="165" cy="138" rx="16" ry="10" fill="url(#lz-leg)" stroke="#0e3a0e" stroke-width="0.6"/>
                </g>
                <!-- back foot with toes -->
                <path d="M 152,148 q -4,3 -10,2  M 154,150 q -4,5 -3,9  M 158,152 q 0,5 3,8  M 163,151 q 3,5 7,6"
                      stroke="#0e3a0e" stroke-width="1.6" fill="none" stroke-linecap="round"/>

                <!-- ── BODY (main mass) ── -->
                <ellipse cx="120" cy="115" rx="58" ry="22" fill="url(#lz-body)" stroke="#0e3a0e" stroke-width="0.9"/>
                <!-- soft yellow underbelly -->
                <ellipse cx="120" cy="125" rx="42" ry="9" fill="url(#lz-belly)" opacity="0.85"/>
                <!-- top highlight strip -->
                <ellipse cx="115" cy="100" rx="40" ry="4" fill="#ffffff" opacity="0.22"/>
                <!-- yellow spots running along body -->
                <circle cx="90"  cy="108" r="3"   fill="#ffd54f"/>
                <circle cx="108" cy="102" r="3.2" fill="#ffd54f"/>
                <circle cx="128" cy="100" r="3"   fill="#ffd54f"/>
                <circle cx="148" cy="105" r="2.8" fill="#ffd54f"/>
                <circle cx="167" cy="110" r="2.6" fill="#ffd54f"/>
                <circle cx="100" cy="118" r="2"   fill="#ffd54f" opacity="0.7"/>
                <circle cx="138" cy="120" r="2"   fill="#ffd54f" opacity="0.7"/>

                <!-- ── FRONT LEG ── -->
                <g transform="rotate(14 80 138)">
                  <ellipse cx="80" cy="138" rx="16" ry="10" fill="url(#lz-leg)" stroke="#0e3a0e" stroke-width="0.6"/>
                </g>
                <!-- front foot with toes splayed -->
                <path d="M 67,148 q -5,2 -10,0  M 68,150 q -5,5 -4,9  M 73,152 q -1,5 2,9  M 78,152 q 2,5 6,7"
                      stroke="#0e3a0e" stroke-width="1.6" fill="none" stroke-linecap="round"/>

                <!-- ── NECK (curve from body to head) ── -->
                <path d="M 75,108 Q 60,90 60,75 Q 70,95 88,108 Z"
                      fill="url(#lz-body)" stroke="#0e3a0e" stroke-width="0.8"/>
                <!-- Yellow throat patch -->
                <ellipse cx="62" cy="88" rx="11" ry="9" fill="url(#lz-belly)" opacity="0.9"/>

                <!-- ── HEAD: big rounded shape ── -->
                <ellipse cx="55" cy="62" rx="38" ry="32" fill="url(#lz-head)" stroke="#0e3a0e" stroke-width="0.9"/>
                <!-- Top highlight on the head -->
                <ellipse cx="48" cy="44" rx="22" ry="6" fill="#ffffff" opacity="0.28"/>

                <!-- ── BIG TOOTHY GRIN ── -->
                <!-- Outer mouth curve -->
                <path d="M 22,68 Q 55,92 88,68" stroke="#0e3a0e" stroke-width="1.6" fill="none" stroke-linecap="round"/>
                <!-- Mouth interior (white) -->
                <path d="M 26,69 Q 55,86 84,69 Q 55,74 26,69 Z" fill="#ffffff" stroke="#0e3a0e" stroke-width="0.7"/>
                <!-- Tooth dividers -->
                <line x1="34" y1="71" x2="34" y2="76" stroke="#bbbbbb" stroke-width="0.6"/>
                <line x1="42" y1="74" x2="42" y2="79" stroke="#bbbbbb" stroke-width="0.6"/>
                <line x1="50" y1="76" x2="50" y2="81" stroke="#bbbbbb" stroke-width="0.6"/>
                <line x1="60" y1="76" x2="60" y2="81" stroke="#bbbbbb" stroke-width="0.6"/>
                <line x1="68" y1="74" x2="68" y2="79" stroke="#bbbbbb" stroke-width="0.6"/>
                <line x1="76" y1="71" x2="76" y2="76" stroke="#bbbbbb" stroke-width="0.6"/>
                <!-- Cheek pink blush -->
                <ellipse cx="22" cy="64" rx="5" ry="3" fill="#ff9aa6" opacity="0.45"/>
                <ellipse cx="86" cy="64" rx="5" ry="3" fill="#ff9aa6" opacity="0.45"/>
                <!-- Tongue (flicks out) -->
                <path id="lz-tongue-path"
                      d="M 55,84 q -2,8 -8,10 q 4,2 8,0 q 4,2 8,0 q -6,-2 -8,-10 z"
                      fill="url(#lz-tongue)" opacity="0"/>

                <!-- ── BIG BULGING EYES on top of head ── -->
                <!-- Left eye (front-left) -->
                <ellipse cx="32" cy="32" rx="15" ry="16" fill="url(#lz-head)" stroke="#0e3a0e" stroke-width="0.9"/>
                <ellipse cx="34" cy="34" rx="11" ry="12" fill="url(#lz-eye)"/>
                <g id="lz-pupil-l">
                  <ellipse cx="37" cy="36" rx="3.4" ry="6.5" fill="#000"/>
                  <circle  cx="38.5" cy="32" r="2"  fill="#ffffff"/>
                </g>
                <ellipse id="lz-lid-l" cx="34" cy="34" rx="12" ry="13"
                         fill="#3a8a25" stroke="#0e3a0e" stroke-width="0.7"/>

                <!-- Right eye (back, slightly higher and bigger) -->
                <ellipse cx="68" cy="26" rx="16" ry="17" fill="url(#lz-head)" stroke="#0e3a0e" stroke-width="0.9"/>
                <ellipse cx="70" cy="28" rx="12" ry="13" fill="url(#lz-eye)"/>
                <g id="lz-pupil-r">
                  <ellipse cx="73" cy="30" rx="3.6" ry="7" fill="#000"/>
                  <circle  cx="74.5" cy="26" r="2.2" fill="#ffffff"/>
                </g>
                <ellipse id="lz-lid-r" cx="70" cy="28" rx="13" ry="14"
                         fill="#3a8a25" stroke="#0e3a0e" stroke-width="0.7"/>

                <!-- Nostril -->
                <circle cx="14" cy="58" r="1.4" fill="#0e3a0e"/>
              </svg>
            </div>
        `;
        doc.body.appendChild(wrap);

        const css = doc.createElement('style');
        css.textContent = `
            #lenny-wrap {
                position: fixed; right: 90px; bottom: 90px;
                z-index: 99999; display: flex; align-items: flex-end; gap: 10px;
                pointer-events: none;
                perspective: 600px;
            }
            #lenny-char {
                position: relative;
                pointer-events: auto;
                line-height: 0;
                cursor: pointer; user-select: none;
                /* Layered shadows for depth and ground contact */
                filter:
                    drop-shadow(0 6px 4px rgba(0,0,0,0.5))
                    drop-shadow(0 14px 10px rgba(0,0,0,0.35))
                    drop-shadow(0 0 14px rgba(143,216,240,0.35));
                transform-style: preserve-3d;
                transform-origin: 50% 92%;
                transition: transform 0.18s ease;
                animation: lenny-idle 4.2s ease-in-out infinite;
                will-change: transform;
            }
            /* Tail wag — pivot at the tail/body joint (175,118) */
            #lz-tail-grp {
                transform-origin: 175px 118px;
                transform-box: view-box;
                animation: lz-tail 1.6s ease-in-out infinite;
            }
            @keyframes lz-tail {
                0%, 100% { transform: rotate(-10deg); }
                50%      { transform: rotate(12deg);  }
            }
            /* Tongue flicks down out of the smile every few seconds */
            #lz-tongue-path {
                transform-origin: 55px 84px;
                transform-box: view-box;
                transform: scaleY(0);
                animation: lz-tongue 4.2s ease-in-out infinite;
            }
            @keyframes lz-tongue {
                0%, 78%, 100% { opacity: 0; transform: scaleY(0); }
                82%           { opacity: 1; transform: scaleY(1); }
                90%           { opacity: 1; transform: scaleY(0.9); }
                95%           { opacity: 0; transform: scaleY(0.4); }
            }
            /* Blink: both lids scale open/closed (default = scaleY 0 = open) */
            #lz-lid-l {
                transform-origin: 34px 34px;
                transform-box: view-box;
                transform: scaleY(0);
                animation: lz-blink 5.2s ease-in-out infinite;
            }
            #lz-lid-r {
                transform-origin: 70px 28px;
                transform-box: view-box;
                transform: scaleY(0);
                animation: lz-blink 5.2s ease-in-out infinite;
            }
            @keyframes lz-blink {
                0%, 92%, 100% { transform: scaleY(0); }
                95%           { transform: scaleY(1); }
            }
            #lenny-char::after {
                /* Soft elliptical ground shadow that pulses with the bob */
                content: '';
                position: absolute;
                left: 50%; bottom: -10px;
                width: 56px; height: 10px;
                transform: translateX(-50%);
                background: radial-gradient(ellipse at center,
                            rgba(0,0,0,0.55) 0%, rgba(0,0,0,0) 70%);
                border-radius: 50%;
                animation: lenny-shadow 4.2s ease-in-out infinite;
                pointer-events: none;
            }
            #lenny-char:hover {
                transform: scale(1.14) rotateY(-18deg) rotateX(6deg);
            }
            #lenny-char.jump  { animation: lenny-jump 1.05s cubic-bezier(.25,.9,.35,1.1); }
            #lenny-char.dance { animation: lenny-dance 2.4s ease-in-out 2; }
            @keyframes lenny-idle {
                0%   { transform: translateY(0)    rotateY(0deg)    rotateX(0deg)   rotateZ(0deg); }
                20%  { transform: translateY(-3px) rotateY(20deg)   rotateX(-2deg)  rotateZ(2deg); }
                50%  { transform: translateY(-8px) rotateY(0deg)    rotateX(-4deg)  rotateZ(0deg); }
                80%  { transform: translateY(-3px) rotateY(-20deg)  rotateX(-2deg)  rotateZ(-2deg); }
                100% { transform: translateY(0)    rotateY(0deg)    rotateX(0deg)   rotateZ(0deg); }
            }
            @keyframes lenny-shadow {
                0%, 100% { opacity: 0.65; transform: translateX(-50%) scale(1); }
                50%      { opacity: 0.35; transform: translateX(-50%) scale(0.7); }
            }
            @keyframes lenny-jump {
                0%   { transform: translateY(0)     rotateY(0)     rotateZ(0)     scale(1); }
                20%  { transform: translateY(-65px) rotateY(180deg) rotateZ(180deg) scale(1.12); }
                50%  { transform: translateY(-110px) rotateY(360deg) rotateZ(540deg) scale(1.18); }
                80%  { transform: translateY(-25px) rotateY(540deg) rotateZ(900deg) scale(1.06); }
                100% { transform: translateY(0)     rotateY(720deg) rotateZ(1080deg) scale(1); }
            }
            /* Boogie! Triggered on tab change, runs for ~5s (2 loops × 2.4s).
               Bounces, sways side to side, hip-twists, head-nods. */
            @keyframes lenny-dance {
                0%   { transform: translate( 0px, 0)     rotateY(0)    rotateZ(0)    scale(1); }
                10%  { transform: translate(-8px, -14px) rotateY(-22deg) rotateZ(-6deg) scale(1.05); }
                25%  { transform: translate(-14px, 0)    rotateY(-30deg) rotateZ(2deg)  scale(1); }
                40%  { transform: translate(-6px, -18px) rotateY(0deg)   rotateZ(8deg)  scale(1.08); }
                55%  { transform: translate( 6px, 0)     rotateY(30deg)  rotateZ(2deg)  scale(1); }
                70%  { transform: translate(14px, -14px) rotateY(22deg)  rotateZ(-6deg) scale(1.05); }
                85%  { transform: translate( 6px, 0)     rotateY(0)      rotateZ(0)    scale(1); }
                100% { transform: translate( 0px, 0)     rotateY(0)      rotateZ(0)    scale(1); }
            }
            #lenny-bubble {
                pointer-events: auto;
                max-width: 240px;
                background: linear-gradient(135deg, #fffbe6 0%, #fff3b0 100%);
                color: #3a2800;
                padding: 12px 30px 12px 14px;
                border-radius: 14px;
                border: 2px solid #FFD54F;
                box-shadow: 0 6px 22px rgba(0,0,0,0.45),
                            0 0 0 2px rgba(255,213,79,0.25);
                font-family: 'Outfit', -apple-system, sans-serif;
                font-size: 13px; line-height: 1.45; font-weight: 500;
                position: relative;
                opacity: 0; transform: translateY(8px) scale(0.9);
                transition: opacity 0.25s ease, transform 0.25s ease;
                pointer-events: none;
            }
            #lenny-bubble.show {
                opacity: 1; transform: translateY(0) scale(1);
                pointer-events: auto;
            }
            #lenny-bubble::after {
                content: ''; position: absolute;
                right: -10px; bottom: 14px;
                width: 0; height: 0;
                border: 8px solid transparent;
                border-left-color: #FFD54F;
            }
            #lenny-close {
                position: absolute; top: 4px; right: 6px;
                background: none; border: none; cursor: pointer;
                color: #6b4a00; font-size: 18px; line-height: 1;
                padding: 2px 6px; border-radius: 50%;
                font-weight: 700;
            }
            #lenny-close:hover { background: rgba(0,0,0,0.08); }
        `;
        doc.head.appendChild(css);

        const charEl   = doc.getElementById('lenny-char');
        const bubble   = doc.getElementById('lenny-bubble');
        const textEl   = doc.getElementById('lenny-text');
        const closeBtn = doc.getElementById('lenny-close');

        let tipIdx = 0;
        let autoTimer = null;
        let bubbleVisible = false;

        function showTip(idx) {
            textEl.textContent = TIPS[idx % TIPS.length];
            bubble.classList.add('show');
            bubbleVisible = true;
            clearTimeout(autoTimer);
            autoTimer = setTimeout(hideBubble, 9000);
        }
        function hideBubble() {
            bubble.classList.remove('show');
            bubbleVisible = false;
        }
        function nextTip() {
            tipIdx = (tipIdx + 1) % TIPS.length;
            showTip(tipIdx);
        }

        charEl.addEventListener('click', () => {
            charEl.classList.remove('jump');
            // force reflow so animation re-triggers
            void charEl.offsetWidth;
            charEl.classList.add('jump');
            nextTip();
        });
        closeBtn.addEventListener('click', (e) => {
            e.stopPropagation();
            hideBubble();
        });

        // ── Dance on tab change ──
        function dance() {
            charEl.classList.remove('dance', 'jump');
            void charEl.offsetWidth;          // re-trigger animation
            charEl.classList.add('dance');
            setTimeout(() => charEl.classList.remove('dance'), 5200);
        }
        function hookTabsForDance() {
            const tabs = doc.querySelectorAll('.stTabs [data-baseweb="tab"]');
            if (tabs.length === 0) return false;
            tabs.forEach(t => {
                if (t.__lennyHooked) return;
                t.__lennyHooked = true;
                t.addEventListener('click', () => {
                    if (t.getAttribute('aria-selected') === 'true') return;
                    dance();
                });
            });
            return true;
        }
        if (!hookTabsForDance()) {
            const obs = new MutationObserver(() => {
                if (hookTabsForDance()) obs.disconnect();
            });
            obs.observe(doc.body, { childList: true, subtree: true });
        }

        // ── Eye-gaze tracking ──
        const svgEl   = doc.getElementById('lenny-svg');
        const pupilL  = doc.getElementById('lz-pupil-l');
        const pupilR  = doc.getElementById('lz-pupil-r');
        // Centers are in SVG-coord space (viewBox 260×180)
        const eyeCenters = [
            { x: 37, y: 36, el: pupilL, max: 4 },
            { x: 73, y: 30, el: pupilR, max: 4 },
        ];
        function updateGaze(mx, my) {
            if (!svgEl) return;
            const r = svgEl.getBoundingClientRect();
            const sx = r.width  / 260;
            const sy = r.height / 180;
            for (const eye of eyeCenters) {
                const cx = r.left + eye.x * sx;
                const cy = r.top  + eye.y * sy;
                const dx = mx - cx, dy = my - cy;
                const d  = Math.hypot(dx, dy) || 1;
                const ramp = Math.min(d, 140) / 140; // smoother near eye
                const ox = (dx / d) * eye.max * ramp;
                const oy = (dy / d) * eye.max * ramp;
                eye.el.setAttribute('transform', `translate(${ox} ${oy})`);
            }
        }
        doc.addEventListener('mousemove', (e) => {
            if (dragging) return;
            updateGaze(e.clientX, e.clientY);
        });

        // ── Drag & throw with physics ──
        let dragging = false;
        let throwing = false;
        let dragOffX = 0, dragOffY = 0;
        let history  = [];

        function pinToLeftTop() {
            // Convert from right/bottom anchoring to left/top so we can move freely
            const r = wrap.getBoundingClientRect();
            wrap.style.left   = r.left + 'px';
            wrap.style.top    = r.top  + 'px';
            wrap.style.right  = 'auto';
            wrap.style.bottom = 'auto';
        }

        charEl.addEventListener('mousedown', (e) => {
            // Only left-button; ignore clicks on the close button etc
            if (e.button !== 0) return;
            e.preventDefault();
            // Cancel any in-flight throw
            throwing = false;
            charEl.classList.remove('jump', 'dance');
            charEl.style.animation = 'none'; // pause idle bob
            charEl.style.transform = '';

            pinToLeftTop();
            const r = wrap.getBoundingClientRect();
            dragOffX = e.clientX - r.left;
            dragOffY = e.clientY - r.top;
            dragging = true;
            charEl.style.cursor = 'grabbing';
            hideBubble();
            history = [{ x: e.clientX, y: e.clientY, t: performance.now() }];
        });

        doc.addEventListener('mousemove', (e) => {
            if (!dragging) return;
            wrap.style.left = (e.clientX - dragOffX) + 'px';
            wrap.style.top  = (e.clientY - dragOffY) + 'px';
            history.push({ x: e.clientX, y: e.clientY, t: performance.now() });
            if (history.length > 6) history.shift();
        });

        doc.addEventListener('mouseup', (e) => {
            if (!dragging) return;
            dragging = false;
            charEl.style.cursor = 'pointer';
            // Distance moved since mousedown — short distance = treat as click
            const a = history[0];
            const b = history[history.length - 1];
            const dist = Math.hypot(b.x - a.x, b.y - a.y);
            if (dist < 6) {
                // Click, not drag. Restore idle animation; existing click handler fires next.
                charEl.style.animation = '';
                return;
            }
            // Suppress the upcoming click event (Lenny's onclick = jump+tip)
            const swallow = (ev) => { ev.stopPropagation(); ev.preventDefault(); };
            doc.addEventListener('click', swallow, { capture: true, once: true });
            // Compute toss velocity from recent mouse history
            const dt = Math.max(b.t - a.t, 16);
            const vx = (b.x - a.x) / dt * 16; // px per frame at ~60fps
            const vy = (b.y - a.y) / dt * 16;
            startThrow(vx, vy);
        });

        function startThrow(vx, vy) {
            throwing = true;
            const r0 = wrap.getBoundingClientRect();
            let x = r0.left, y = r0.top;
            let rot = 0;
            // Spin proportional to throw speed, with a min so a gentle drop still tumbles
            const speed = Math.hypot(vx, vy);
            let vr = (vx >= 0 ? 1 : -1) * Math.max(speed * 0.8, 4);
            const gravity = 0.7;
            const bounce  = 0.45;
            const friction = 0.985;
            let stillFrames = 0;

            function step() {
                if (!throwing) return;
                vy += gravity;
                x  += vx;
                y  += vy;
                rot += vr;

                const w = wrap.offsetWidth;
                const h = wrap.offsetHeight;
                const maxX = (window.innerWidth || doc.documentElement.clientWidth) - w - 4;
                const maxY = (window.innerHeight || doc.documentElement.clientHeight) - h - 4;

                // Walls
                if (x < 4)    { x = 4;    vx = -vx * bounce; vr = -vr * 0.7; }
                if (x > maxX) { x = maxX; vx = -vx * bounce; vr = -vr * 0.7; }
                // Ceiling
                if (y < 4)    { y = 4;    vy = -vy * bounce; }

                // Ground
                if (y >= maxY) {
                    y = maxY;
                    vy = -vy * bounce;
                    vx *= friction;
                    vr *= 0.7;
                    if (Math.abs(vy) < 1.6 && Math.abs(vx) < 0.4) {
                        stillFrames++;
                    } else {
                        stillFrames = 0;
                    }
                }

                wrap.style.left = x + 'px';
                wrap.style.top  = y + 'px';
                charEl.style.transform = `rotate(${rot}deg)`;

                if (stillFrames > 5) {
                    // Settle: snap upright, restore idle animation
                    throwing = false;
                    charEl.style.transform  = '';
                    charEl.style.animation  = '';
                    return;
                }
                requestAnimationFrame(step);
            }
            requestAnimationFrame(step);
        }

        // Greet on first load after a short pause
        setTimeout(() => showTip(0), 1500);
        // Periodic gentle reminder if hidden
        setInterval(() => {
            if (!bubbleVisible) showTip((tipIdx + 1) % TIPS.length);
        }, 45000);
    })();
    </script>
    """, height=0)


render_lenny_lizard()


# ─────────────────────────────────────────────────────────────
# 🥚 EASTER EGGS — Konami codes, click counters, date triggers,
#                  Lenny variants, hidden routes, Possum cameo
# ─────────────────────────────────────────────────────────────
def render_easter_eggs():
    import streamlit.components.v1 as _components
    _components.html(r"""
    <script>
    (function() {
        const doc = window.parent.document;
        if (doc.__eeInit) return;
        doc.__eeInit = true;

        // ─── styles ───
        const css = doc.createElement('style');
        css.textContent = `
            @keyframes ee-fw   { 0%{transform:translate(0,0) scale(.3);opacity:1} 100%{transform:translate(var(--dx),var(--dy)) scale(1);opacity:0} }
            .ee-firework { position:fixed; width:9px; height:9px; border-radius:50%; pointer-events:none; z-index:99980;
                           animation: ee-fw 1.3s ease-out forwards; }
            @keyframes ee-pop  { 0%{transform:translate(-50%,-50%) scale(0) rotate(-12deg);opacity:0}
                                 25%{transform:translate(-50%,-50%) scale(1.18) rotate(0);opacity:1}
                                 80%{transform:translate(-50%,-50%) scale(1) rotate(0);opacity:1}
                                 100%{transform:translate(-50%,-50%) scale(0.65) rotate(8deg);opacity:0} }
            .ee-banner { position:fixed; pointer-events:none; z-index:99998;
                         background:linear-gradient(135deg,#FFE98A,#FFA500); color:#3a2800;
                         font-family:'Shrikhand',cursive; font-size:34px; padding:14px 26px;
                         border-radius:18px; border:3px solid #FFD54F;
                         box-shadow:0 8px 30px rgba(0,0,0,0.5);
                         animation: ee-pop 2.2s ease-out forwards;
                         text-shadow:0 2px 0 rgba(0,0,0,0.2); white-space:nowrap; }
            @keyframes ee-meat-cross { from{transform:translateX(-120px)} to{transform:translateX(calc(100vw + 120px))} }
            @keyframes ee-meat-bob   { 0%,100%{transform:translateY(0)} 50%{transform:translateY(-26px)} }
            .ee-meat-strip { position:fixed; bottom:25%; left:0; pointer-events:none; z-index:99970;
                             font-size:54px; display:flex; gap:6px; will-change:transform;
                             animation: ee-meat-cross 5.8s linear forwards; }
            .ee-meat-strip span { display:inline-block; animation: ee-meat-bob .55s ease-in-out infinite; }
            .ee-gh { position:fixed; inset:0; z-index:99990; pointer-events:none; opacity:0;
                     background: radial-gradient(ellipse at center, rgba(80,40,0,.86), rgba(20,10,0,.98));
                     color:#FFE98A; font-family:'Shrikhand',cursive; font-size:48px;
                     display:flex; align-items:center; justify-content:center;
                     text-align:center; line-height:1.15;
                     transition: opacity .6s ease;
                     text-shadow:0 0 22px #FFA500, 0 0 6px #ff5500; }
            .ee-gh.show { opacity:1; }
            @keyframes ee-possum-walk { from{transform:translateX(-90px)} to{transform:translateX(calc(100vw + 90px))} }
            .ee-possum { position:fixed; bottom:6px; left:0; pointer-events:none; z-index:99970;
                         font-size:42px; animation: ee-possum-walk 14s linear forwards;
                         filter: drop-shadow(0 4px 4px rgba(0,0,0,.5)); }
            @keyframes ee-zzz { from{transform:translateY(0) rotate(0); opacity:1}
                                to  {transform:translateY(-90px) rotate(22deg); opacity:0} }
            .ee-zzz { position:fixed; pointer-events:none; z-index:99999; font-size:30px;
                      color:#fff; font-family:'Shrikhand',cursive; will-change:transform;
                      animation: ee-zzz 2.6s ease-out forwards;
                      text-shadow:0 0 8px rgba(255,255,255,.6); }
            .ee-shades { position:absolute; top:8%; left:14%; font-size:34px; pointer-events:none;
                         filter: drop-shadow(0 2px 1px rgba(0,0,0,.5)); transform:rotate(-4deg); }
            #ee-debug { position:fixed; top:60px; right:20px; z-index:99999;
                        background:rgba(20,20,36,0.96); color:#8fd8f0;
                        padding:14px 18px; border-radius:10px;
                        border:1px solid rgba(255,255,255,0.15);
                        font-family:monospace; font-size:11px;
                        max-width:340px; line-height:1.55;
                        box-shadow:0 8px 28px rgba(0,0,0,.55); }
            @keyframes ee-spin { to { transform: rotate(720deg); } }
            .ee-spinning { animation: ee-spin 1.8s cubic-bezier(.3,.9,.3,1) !important;
                           display:inline-block; transform-origin:50% 50%; }
            @keyframes ee-confetti { to { transform: translateY(105vh) rotate(720deg); } }
            .ee-confetti { position:fixed; top:-20px; width:8px; height:14px;
                           pointer-events:none; z-index:99970;
                           animation: ee-confetti 3.2s linear forwards; }
            .ee-halloween-tint { position:fixed; inset:0; pointer-events:none; z-index:1;
                                 background: radial-gradient(circle at 70% 30%, rgba(255,80,0,0.10), transparent 55%),
                                             radial-gradient(circle at 20% 80%, rgba(120,0,180,0.08), transparent 55%);
                                 mix-blend-mode: screen; }
            #ee-nye { position:fixed; top:10px; left:50%; transform:translateX(-50%); z-index:99999;
                      background:linear-gradient(135deg,#2a2a4a,#1a1a2e); color:#FFE98A;
                      padding:8px 18px; border-radius:22px; border:1px solid #FFD54F;
                      font-family:'Shrikhand',cursive; font-size:14px;
                      box-shadow:0 6px 20px rgba(0,0,0,.45); }
            /* Shoe flies at the camera */
            @keyframes ee-shoe-fly {
                0%   { transform: translate(-50%,-50%) translate(40vw, -45vh) scale(0.05) rotate(-30deg); opacity: 0; }
                15%  { opacity: 1; }
                70%  { transform: translate(-50%,-50%) translate(0, 0) scale(7) rotate(720deg); opacity: 1; }
                85%  { transform: translate(-50%,-50%) translate(0, 0) scale(9) rotate(810deg); opacity: 1; }
                100% { transform: translate(-50%,-50%) translate(0, 0) scale(11) rotate(900deg); opacity: 0; }
            }
            .ee-shoe { position:fixed; left:50%; top:50%; pointer-events:none; z-index:99999;
                       font-size:60px; will-change: transform;
                       filter: drop-shadow(0 8px 14px rgba(0,0,0,0.65));
                       animation: ee-shoe-fly 1.4s cubic-bezier(.55,.05,.85,.5) forwards; }
            /* Whole-page shake when shoe "hits" */
            @keyframes ee-shake {
                0%, 100% { transform: translate(0,0); }
                10% { transform: translate(-6px, 4px); }
                20% { transform: translate(7px, -3px); }
                30% { transform: translate(-5px, -5px); }
                40% { transform: translate(8px, 2px); }
                50% { transform: translate(-4px, 6px); }
                60% { transform: translate(5px, -4px); }
                70% { transform: translate(-7px, 3px); }
                80% { transform: translate(4px, 4px); }
                90% { transform: translate(-3px, -2px); }
            }
            .ee-shaking { animation: ee-shake 0.55s cubic-bezier(.36,.07,.19,.97); }
            /* Crack overlay flash on impact */
            @keyframes ee-crack { 0%{opacity:0} 12%{opacity:1} 100%{opacity:0} }
            .ee-crack { position:fixed; inset:0; pointer-events:none; z-index:99998;
                        background: radial-gradient(ellipse at 50% 50%, rgba(255,255,255,.25) 0%, transparent 35%);
                        animation: ee-crack 0.6s ease-out forwards; }
            @keyframes ee-donut-roll { from{transform:rotate(0)} to{transform:rotate(360deg)} }
            .ee-donut-fall { position:fixed; pointer-events:none; z-index:99970;
                             font-size:48px; animation: ee-donut-roll 1s linear infinite,
                             ee-fall 4s linear forwards; }
            @keyframes ee-fall { to { transform: translateY(110vh) rotate(720deg); } }
        `;
        doc.head.appendChild(css);

        // ─── helpers ───
        function spawnBanner(text, top, fontSize) {
            const b = doc.createElement('div');
            b.className = 'ee-banner';
            b.textContent = text;
            b.style.left = '50%';
            b.style.top  = (top || '40%');
            if (fontSize) b.style.fontSize = fontSize;
            doc.body.appendChild(b);
            setTimeout(() => b.remove(), 2300);
        }

        function fireworks(cx, cy, count) {
            const colors = ['#FFD54F','#ff5fa2','#8fd8f0','#90EE90','#ce93d8','#ff8a65'];
            const N = count || 24;
            for (let i = 0; i < N; i++) {
                const p = doc.createElement('div');
                p.className = 'ee-firework';
                const a = (i / N) * 2 * Math.PI + Math.random()*0.2;
                const r = 70 + Math.random() * 90;
                p.style.left = cx + 'px';
                p.style.top  = cy + 'px';
                p.style.background = colors[i % colors.length];
                p.style.setProperty('--dx', Math.cos(a) * r + 'px');
                p.style.setProperty('--dy', Math.sin(a) * r + 'px');
                doc.body.appendChild(p);
                setTimeout(() => p.remove(), 1400);
            }
        }

        // ─── Konami / phrase triggers ───
        const TRIGGERS = {
            yem: () => {
                for (let i=0; i<5; i++) setTimeout(() => {
                    fireworks(Math.random()*window.innerWidth, 100 + Math.random()*window.innerHeight*0.5);
                }, i*230);
                spawnBanner('YOU ENJOY MYSELF!', '38%');
            },
            wilson: () => {
                for (let i=0; i<6; i++) setTimeout(() => {
                    const w = doc.createElement('div');
                    w.className = 'ee-banner';
                    w.textContent = 'WILLL-SONNN!';
                    w.style.left = (15 + Math.random()*70) + '%';
                    w.style.top  = (15 + Math.random()*60) + '%';
                    doc.body.appendChild(w);
                    setTimeout(() => w.remove(), 2200);
                }, i*340);
            },
            gamehendge: () => {
                const o = doc.createElement('div');
                o.className = 'ee-gh';
                o.innerHTML = '🏰 Welcome to<br>GAMEHENDGE 🐉';
                doc.body.appendChild(o);
                requestAnimationFrame(() => o.classList.add('show'));
                setTimeout(() => { o.classList.remove('show'); setTimeout(()=>o.remove(),700); }, 7000);
            },
            meatstick: () => {
                const strip = doc.createElement('div');
                strip.className = 'ee-meat-strip';
                for (let i=0; i<14; i++) {
                    const s = doc.createElement('span');
                    s.textContent = '🌭';
                    s.style.animationDelay = (i*0.07) + 's';
                    strip.appendChild(s);
                }
                doc.body.appendChild(strip);
                setTimeout(() => strip.remove(), 6300);
            },
            fuego: () => {
                for (let i=0; i<26; i++) setTimeout(() => {
                    const f = doc.createElement('div');
                    f.style.cssText = 'position:fixed;pointer-events:none;z-index:99980;font-size:36px';
                    f.textContent = '🔥';
                    f.style.left = Math.random()*window.innerWidth + 'px';
                    f.style.top  = window.innerHeight + 'px';
                    f.style.transition = 'transform 1.6s ease-out, opacity 1.6s ease-out';
                    doc.body.appendChild(f);
                    requestAnimationFrame(() => {
                        f.style.transform = 'translateY(-' + (60 + Math.random()*window.innerHeight*0.7) + 'px)';
                        f.style.opacity = '0';
                    });
                    setTimeout(() => f.remove(), 1800);
                }, i*55);
                spawnBanner('🔥 FUEGO 🔥', '40%');
            },
            tweezer: () => {
                doc.body.style.transition = 'filter 0.8s ease';
                doc.body.style.filter = 'hue-rotate(180deg) saturate(2)';
                setTimeout(() => { doc.body.style.filter = ''; }, 2400);
                spawnBanner('🌀 TWEEZER REPRISE 🌀', '40%');
            },
            divided: () => {
                spawnBanner('DIVIDED SKY ☁️', '40%', '28px');
            },
            cavern: () => {
                // 👟 "whatever you do, take care of your shoes" — shoe flies at the screen
                const shoe = doc.createElement('div');
                shoe.className = 'ee-shoe';
                shoe.textContent = '👟';
                doc.body.appendChild(shoe);
                // After ~1s, the shoe "hits": shake the page + flash + remove
                setTimeout(() => {
                    const flash = doc.createElement('div');
                    flash.className = 'ee-crack';
                    doc.body.appendChild(flash);
                    setTimeout(() => flash.remove(), 600);
                    const root = doc.querySelector('.stApp') || doc.body;
                    root.classList.add('ee-shaking');
                    setTimeout(() => root.classList.remove('ee-shaking'), 600);
                }, 980);
                setTimeout(() => shoe.remove(), 1500);
                setTimeout(() => spawnBanner('whatever you do, take care of your shoes', '14%', '20px'), 1100);
            }
        };
        let keybuf = '';
        doc.addEventListener('keydown', (e) => {
            // Skip when typing into inputs
            const tgt = e.target;
            if (tgt && (tgt.tagName === 'INPUT' || tgt.tagName === 'TEXTAREA' || tgt.isContentEditable)) return;
            if (!/^[a-zA-Z]$/.test(e.key)) return;
            keybuf = (keybuf + e.key.toLowerCase()).slice(-24);
            for (const k in TRIGGERS) {
                if (keybuf.endsWith(k)) { TRIGGERS[k](); keybuf=''; break; }
            }
        });

        // ─── Possum walk-on (1-in-200 page loads) ───
        if (Math.random() < 1/200) {
            setTimeout(() => {
                const p = doc.createElement('div');
                p.className = 'ee-possum';
                p.textContent = '🐀';
                doc.body.appendChild(p);
                setTimeout(() => p.remove(), 14500);
            }, 5000 + Math.random()*25000);
        }

        // ─── Lenny: 60s idle → sleep z's ───
        let lastTouch = Date.now();
        ['mousemove','keydown','click','scroll','touchstart'].forEach(ev =>
            doc.addEventListener(ev, () => lastTouch = Date.now(), { passive:true }));
        setInterval(() => {
            if (Date.now() - lastTouch < 60000) return;
            const lenny = doc.getElementById('lenny-char');
            if (!lenny) return;
            const r = lenny.getBoundingClientRect();
            const z = doc.createElement('div');
            z.className = 'ee-zzz';
            z.textContent = ['z','Z','zz'][Math.floor(Math.random()*3)];
            z.style.left = (r.left + r.width/2 + (Math.random()-0.5)*24) + 'px';
            z.style.top  = (r.top - 10) + 'px';
            doc.body.appendChild(z);
            setTimeout(() => z.remove(), 2700);
        }, 1700);

        // ─── Lenny: 10× rapid click → sunglasses ───
        let lennyClicks = []; let shadesOn = false;
        doc.addEventListener('click', (e) => {
            const ln = doc.getElementById('lenny-char');
            if (!ln || !ln.contains(e.target)) return;
            const now = Date.now();
            lennyClicks = lennyClicks.filter(t => now - t < 5000);
            lennyClicks.push(now);
            if (lennyClicks.length >= 10 && !shadesOn) {
                shadesOn = true;
                const shades = doc.createElement('div');
                shades.className = 'ee-shades';
                shades.textContent = '🕶️';
                ln.appendChild(shades);
            }
        });

        // ─── Title 5× rapid → spin like a vinyl ───
        function hookTitleAndFooter() {
            doc.querySelectorAll('h1').forEach(h => {
                if (h.__eeHooked) return;
                h.__eeHooked = true; let tc = [];
                h.style.cursor = 'pointer';
                h.addEventListener('click', () => {
                    const now = Date.now();
                    tc = tc.filter(t => now - t < 1500);
                    tc.push(now);
                    if (tc.length >= 5) {
                        h.classList.add('ee-spinning');
                        setTimeout(() => h.classList.remove('ee-spinning'), 1900);
                        tc = [];
                    }
                });
            });
            const footer = doc.querySelector('.gj-footer');
            if (footer && !footer.__eeHooked) {
                footer.__eeHooked = true; let ft = []; footer.style.cursor = 'pointer';
                footer.addEventListener('click', () => {
                    const now = Date.now();
                    ft = ft.filter(t => now - t < 1000);
                    ft.push(now);
                    if (ft.length >= 3) {
                        let dbg = doc.getElementById('ee-debug');
                        if (dbg) { dbg.remove(); ft = []; return; }
                        dbg = doc.createElement('div');
                        dbg.id = 'ee-debug';
                        const sel = doc.querySelector('.stTabs [aria-selected="true"]');
                        dbg.innerHTML =
                            '<b style="color:#FFE98A">🐛 GJ DEBUG</b><br>' +
                            'Viewport: ' + innerWidth + '×' + innerHeight + '<br>' +
                            'Active tab: ' + (sel ? sel.textContent : '?') + '<br>' +
                            'Lenny: ' + (doc.getElementById('lenny-char') ? '🟢 alive' : '🔴 missing') + '<br>' +
                            'Date: ' + new Date().toISOString().slice(0,10) + '<br>' +
                            'Phrase triggers: <b>yem · wilson · gamehendge · meatstick · fuego · tweezer · divided · cavern</b><br>' +
                            'URL params: <code>?donut=1</code> turns Lenny into a donut<br>' +
                            '<span style="color:#888">triple-click footer to close</span>';
                        doc.body.appendChild(dbg);
                        ft = [];
                    }
                });
            }
        }
        hookTitleAndFooter();
        new MutationObserver(hookTitleAndFooter).observe(doc.body, { childList:true, subtree:true });

        // ─── Donut emoji hover → falls and rolls ───
        function hookDonuts() {
            doc.body.querySelectorAll('h1,h2,h3,h4,p,div,span').forEach(el => {
                if (el.__eeDonut) return;
                if (!el.textContent || el.textContent.length > 200) return;
                if (!el.textContent.includes('🍩')) return;
                el.__eeDonut = true;
                let hoverStart = 0;
                el.addEventListener('mouseenter', () => { hoverStart = Date.now(); });
                el.addEventListener('mouseleave', () => {
                    if (Date.now() - hoverStart < 1200) return;
                    const r = el.getBoundingClientRect();
                    const d = doc.createElement('div');
                    d.className = 'ee-donut-fall';
                    d.textContent = '🍩';
                    d.style.left = (r.left + r.width/2) + 'px';
                    d.style.top  = (r.top + r.height/2) + 'px';
                    doc.body.appendChild(d);
                    setTimeout(() => d.remove(), 4200);
                });
            });
        }
        hookDonuts();
        new MutationObserver(hookDonuts).observe(doc.body, { childList:true, subtree:true });

        // ─── URL param: ?donut=1 → Lenny becomes a donut ───
        const search = window.parent.location.search || '';
        if (/donut=1/.test(search)) {
            const tryDonut = () => {
                const ln = doc.getElementById('lenny-char');
                if (!ln) return false;
                ln.innerHTML = '<div style="font-size:80px;line-height:1;text-shadow:0 4px 6px rgba(0,0,0,.4)">🍩</div>';
                return true;
            };
            if (!tryDonut()) {
                const obs = new MutationObserver(() => { if (tryDonut()) obs.disconnect(); });
                obs.observe(doc.body, { childList:true, subtree:true });
            }
        }

        // ─── Date-based eggs ───
        const today = new Date();
        const m = today.getMonth() + 1, d = today.getDate();

        // Halloween 10/31
        if (m === 10 && d === 31) {
            const tint = doc.createElement('div');
            tint.className = 'ee-halloween-tint';
            doc.body.appendChild(tint);
            setTimeout(() => spawnBanner('🎃 Spooky Phish! 👻', '12%', '24px'), 1200);
        }

        // Trey's birthday 9/30
        if (m === 9 && d === 30) {
            const colors = ['#FFD54F','#ff5fa2','#8fd8f0','#90EE90','#ce93d8','#ff8a65'];
            for (let i=0; i<90; i++) setTimeout(() => {
                const c = doc.createElement('div');
                c.className = 'ee-confetti';
                c.style.left = Math.random()*100 + 'vw';
                c.style.background = colors[i % colors.length];
                c.style.animationDuration = (2.6 + Math.random()*2) + 's';
                doc.body.appendChild(c);
                setTimeout(() => c.remove(), 5200);
            }, i*70);
            setTimeout(() => spawnBanner('🎂 Happy Birthday Trey 🎂', '14%', '26px'), 200);
        }

        // NYE 12/31 countdown widget
        if (m === 12 && d === 31) {
            const cd = doc.createElement('div');
            cd.id = 'ee-nye';
            doc.body.appendChild(cd);
            const tick = () => {
                const now = new Date();
                const next = new Date(now.getFullYear()+1, 0, 1, 0, 0, 0);
                let s = Math.max(0, Math.floor((next - now) / 1000));
                const h = Math.floor(s/3600); s -= h*3600;
                const mm = Math.floor(s/60);   s -= mm*60;
                cd.textContent = '🎆 NYE ' + String(h).padStart(2,'0') + ':' +
                                 String(mm).padStart(2,'0') + ':' + String(s).padStart(2,'0') + ' 🎆';
            };
            tick(); setInterval(tick, 1000);
        }

        // ─── Antelope counter — counts in localStorage how many predictions a user
        //     has loaded that contained "Run Like an Antelope". Display via debug panel.
        try {
            const KEY = 'gj_antelope_count';
            const setlists = doc.querySelectorAll('table');
            let found = false;
            setlists.forEach(t => {
                if (t.textContent.includes('Run Like an Antelope')) found = true;
            });
            if (found) {
                let n = parseInt(localStorage.getItem(KEY) || '0', 10);
                localStorage.setItem(KEY, String(n + 1));
            }
        } catch(_) {}
    })();
    </script>
    """, height=0)


render_easter_eggs()


tab1, tab2, tab3 = st.tabs(["🎸 City Predictor", "🏟️ Top 50 · Sphere 2026", "🔮 Sphere Predictor"])

# ── Glowstick rain on tab change ─────────────────────────────
# Injected via components.v1.html so the <script> actually runs.
# The JS reaches into the parent document to attach tab-click listeners
# and spawn glowsticks that arc from top to bottom.
import streamlit.components.v1 as _components
_components.html(r"""
<script>
(function() {
    const doc = window.parent.document;
    if (doc.__gjGlowInit) return;
    doc.__gjGlowInit = true;

    // Inject stylesheet into parent
    const style = doc.createElement('style');
    style.textContent = `
    .gj-glow-layer {
        position: fixed; inset: 0; pointer-events: none; z-index: 9998;
        overflow: hidden;
    }
    .gj-glow {
        position: absolute; top: -8vh; font-size: 1.8rem;
        filter: drop-shadow(0 0 8px currentColor) drop-shadow(0 0 14px currentColor);
        will-change: transform, opacity;
    }
    @keyframes gj-glow-arc-L {
        0%   { opacity: 0; transform: translate(0, 0) rotate(0deg); }
        10%  { opacity: 1; }
        100% { opacity: 0; transform: translate(-25vw, 115vh) rotate(540deg); }
    }
    @keyframes gj-glow-arc-R {
        0%   { opacity: 0; transform: translate(0, 0) rotate(0deg); }
        10%  { opacity: 1; }
        100% { opacity: 0; transform: translate(25vw, 115vh) rotate(-540deg); }
    }
    @keyframes gj-glow-arc-C {
        0%   { opacity: 0; transform: translate(0, 0) rotate(0deg) scale(0.9); }
        10%  { opacity: 1; }
        100% { opacity: 0; transform: translate(0, 115vh) rotate(360deg) scale(1.1); }
    }
    @media (prefers-reduced-motion: reduce) { .gj-glow { display: none; } }
    `;
    doc.head.appendChild(style);

    const colors = ['#ff3b6b', '#3bffb0', '#3bbfff', '#ffe83b', '#c13bff', '#ff8c3b'];
    const glyphs = ['\u2728', '\uD83D\uDD6F\uFE0F', '\u2728', '\uD83C\uDF1F', '\uD83D\uDD6F\uFE0F', '\u2728'];

    function spawnGlowsticks() {
        const layer = doc.createElement('div');
        layer.className = 'gj-glow-layer';
        doc.body.appendChild(layer);

        const N = 28;
        for (let i = 0; i < N; i++) {
            const s = doc.createElement('span');
            s.className = 'gj-glow';
            const color = colors[i % colors.length];
            s.style.color = color;
            // 🥚 1-in-50 easter egg: a random glowstick is replaced by a chess piece
            s.textContent = (Math.random() < 1/50)
                ? ['♛','♞','♜','♝','♚','♟'][Math.floor(Math.random()*6)]
                : glyphs[i % glyphs.length];
            s.style.left = (2 + (i * 96 / N) + (Math.random() * 4 - 2)) + 'vw';
            const dur = 2.2 + Math.random() * 1.8;
            const delay = Math.random() * 0.9;
            const arc = ['L', 'R', 'C'][i % 3];
            s.style.animation = `gj-glow-arc-${arc} ${dur}s cubic-bezier(0.35, 0.1, 0.6, 1) ${delay}s forwards`;
            s.style.fontSize = (1.4 + Math.random() * 1.2) + 'rem';
            layer.appendChild(s);
        }
        setTimeout(() => layer.remove(), 5200);
    }

    function hookTabs() {
        const tabs = doc.querySelectorAll('.stTabs [data-baseweb="tab"]');
        if (tabs.length === 0) return false;
        tabs.forEach(t => {
            if (t.__gjHooked) return;
            t.__gjHooked = true;
            t.addEventListener('click', () => {
                // Only spawn if user is actually switching (not re-clicking current tab)
                if (t.getAttribute('aria-selected') === 'true') return;
                spawnGlowsticks();
            });
        });
        return true;
    }

    // Try immediately; retry if tabs aren't mounted yet
    if (!hookTabs()) {
        const obs = new MutationObserver(() => { if (hookTabs()) obs.disconnect(); });
        obs.observe(doc.body, { childList: true, subtree: true });
        setTimeout(() => obs.disconnect(), 10000);
    }
})();
</script>
""", height=0)

# ═══════════════════════════════════════════════════════════
# TAB 1 — Setlist Predictor
# ═══════════════════════════════════════════════════════════
with tab1:
    st.divider()

    # Pre-fill from ?city=... query param
    _url_city = st.query_params.get("city", "")
    city = st.text_input(
        "Enter a city:",
        value=_url_city,
        placeholder="e.g. Albany, Chicago, Noblesville",
        key="city_input",
    )

    if city:
        with st.spinner(f"Generating setlist for {city}..."):
            rows, city_shows, locations = generate_setlist(city)

        if rows is None:
            st.error(f"No shows found for '{city}'. Try a different city name.")
        else:
            st.success(f"Found **{city_shows} shows** in {', '.join(locations)} — setlist has **{len(rows)} songs**")

            # Color-code the table
            tier_colors = {
                "Staple":     "#FFD700",
                "Common":     "#4FC3F7",
                "Occasional": "#A5D6A7",
                "Rare":       "#CE93D8",
            }

            header_cols = ["#", "Song", "Tier", "City Freq", "Gap", "Adj Score", "Role"]
            col_widths_pct = [4, 30, 12, 10, 8, 10, 12]

            header_html = "".join(
                f'<th style="background:#1A1A2E;color:#F0E68C;padding:8px;text-align:center;width:{w}%">{h}</th>'
                for h, w in zip(header_cols, col_widths_pct)
            )

            rows_html = ""
            prev_set = None
            for row in rows:
                cur_set = row.get("Set", "")
                if cur_set != prev_set:
                    banner = {"Set 1":"🎸 SET 1","Set 2":"🎸 SET 2","Encore":"🎤 ENCORE"}.get(cur_set, cur_set)
                    rows_html += (
                        f'<tr><td colspan="7" style="background:#0d0d1e;color:#FFD54F;'
                        f'padding:10px 8px;font-weight:700;letter-spacing:0.1em;'
                        f'font-family:Shrikhand,cursive;font-size:1.05rem">{banner}</td></tr>'
                    )
                    prev_set = cur_set

                is_bustout = row.get("Bust Out", False)
                role = row.get("Role", "")
                is_opener = role == "Opener"
                is_closer = role == "Closer"
                is_encore = cur_set == "Encore"

                if is_bustout:   bg, fg = "#1A2E1A", "#B9F6CA"
                elif is_encore:  bg, fg = "#2a1a3e", "#E1BEE7"
                elif is_closer:  bg, fg = "#2E1A1A", "#F4C2C2"
                elif is_opener:  bg, fg = "#1a2a3e", "#B3E5FC"
                else:            bg, fg = "#1a1a2e" if row["#"] % 2 == 0 else "#16213e", "#EEEEEE"

                tier_color = tier_colors.get(row["Tier"], "#FFFFFF")
                adj_color = "#66BB6A" if row["_adj"] >= 30 else ("#42A5F5" if row["_adj"] >= 25 else fg)
                song_label = f"{row['Song']} ⭐ <span style='color:#B9F6CA;font-size:11px'>BUST OUT</span>" if is_bustout else row['Song']
                role_label = role if role else "—"

                rows_html += f"""
                <tr style="background:{bg};color:{fg}">
                    <td style="text-align:center;padding:6px">{row['#']}</td>
                    <td style="padding:6px;font-weight:bold">{song_label}</td>
                    <td style="text-align:center;padding:6px;color:{tier_color}">{row['Tier']}</td>
                    <td style="text-align:center;padding:6px">{row['City Freq']}</td>
                    <td style="text-align:center;padding:6px">{row['Shows Since Last Played']}</td>
                    <td style="text-align:center;padding:6px;color:{adj_color};font-weight:bold">{row['Adj Score']}</td>
                    <td style="text-align:center;padding:6px">{role_label}</td>
                </tr>"""

            table_html = f"""
            <table style="width:100%;border-collapse:collapse;font-family:Arial;font-size:14px">
                <thead><tr>{header_html}</tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
            """
            st.markdown(table_html, unsafe_allow_html=True)

            st.markdown("""
            <div style="font-size:11px;color:#666;margin-top:8px">
            🟡 Staple &nbsp;|&nbsp; 🔵 Common &nbsp;|&nbsp; 🟢 Occasional &nbsp;|&nbsp; 🟣 Rare &nbsp;|&nbsp;
            <span style="background:#1a2a3e;color:#B3E5FC;padding:1px 4px">Blue = opener</span> &nbsp;|&nbsp;
            <span style="background:#2E1A1A;color:#F4C2C2;padding:1px 4px">Red = closer</span> &nbsp;|&nbsp;
            <span style="background:#2a1a3e;color:#E1BEE7;padding:1px 4px">Purple = encore</span> &nbsp;|&nbsp;
            <span style="background:#1A2E1A;color:#B9F6CA;padding:1px 4px">Green ⭐ = Bust Out (gap > 500)</span>
            </div>
            """, unsafe_allow_html=True)

            st.divider()

            # Highlights
            top = max(rows, key=lambda r: r["_adj"])
            bustouts = [r for r in rows if r.get("Bust Out")]
            encores = [r for r in rows if r.get("Set") == "Encore"]
            s2_closer = next((r for r in rows if r.get("Set")=="Set 2" and r.get("Role")=="Closer"), None)

            st.markdown(f"**Top pick:** {top['Song']} ({top['Adj Score']} adj score) — the most probable song based on city history and gap.")
            if bustouts:
                b = bustouts[0]
                st.markdown(f"**⭐ Bust Out:** {b['Song']} — overdue by {b['_gap']} shows.")
            if s2_closer:
                st.markdown(f"**Set 2 closer:** {s2_closer['Song']}")
            if encores:
                st.markdown(f"**Encore:** {' → '.join(e['Song'] for e in encores)}")

            # Share link
            render_share_box({"city": city}, key="city")

            st.divider()

            # Download button
            xlsx_buf = build_xlsx(rows, city, city_shows)
            st.download_button(
                label="⬇️ Download Spreadsheet (.xlsx)",
                data=xlsx_buf,
                file_name=f"{city.replace(' ', '_')}_Setlist.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.divider()

            # Bingo card
            st.subheader("🎲 Bingo Card")
            st.caption("🟡 From predicted setlist · 🔵 Globally common · 🟣 Rare / uncommon · ★ FREE center")
            bsize_label = st.radio(
                "Card size", ["5×5 Classic", "3×3 Quick"],
                horizontal=True, key=f"bingo_size_{city}",
            )
            bsize = 3 if bsize_label.startswith("3") else 5

            if st.button("Generate Bingo Card"):
                raw = generate_bingo(city)
                if raw:
                    cards = _inject_free_center(raw, bsize)
                    st.session_state["bingo_cards"] = cards
                    st.session_state["bingo_city"]  = city
                    st.session_state["bingo_size"]  = bsize

            if ("bingo_cards" in st.session_state
                    and st.session_state.get("bingo_city") == city
                    and st.session_state.get("bingo_size") == bsize):
                cards = st.session_state["bingo_cards"]
                cat_styles = {
                    "setlist": "background:#2a2a10;color:#F0E68C;border:1px solid #555522",
                    "common":  "background:#0d2235;color:#7ec8e3;border:1px solid #1a4a66",
                    "rare":    "background:#2a0d35;color:#ce93d8;border:1px solid #5a2a6a",
                }
                free_style = ("background:linear-gradient(135deg,#5a3d00,#3a2800);color:#FFF3B0;"
                              "border:1.5px solid #FFD54F;font-family:'Shrikhand',cursive")
                cell_style = ("padding:8px 4px;text-align:center;font-size:12px;font-weight:bold;"
                              "border-radius:6px;min-height:60px;display:flex;align-items:center;"
                              "justify-content:center;word-break:break-word;")

                col_labels = ["P", "H", "I"] if bsize == 3 else ["P", "H", "I", "S", "H"]
                header_cols_b = st.columns(bsize)
                for col, label in zip(header_cols_b, col_labels):
                    col.markdown(f'<div style="text-align:center;font-size:22px;font-weight:bold;color:#F0E68C">{label}</div>', unsafe_allow_html=True)

                for row_i in range(bsize):
                    cols = st.columns(bsize)
                    for col_i, col in enumerate(cols):
                        card = cards[row_i * bsize + col_i]
                        style = free_style if card.get("free") else cat_styles[card["cat"]]
                        col.markdown(
                            f'<div style="{style};{cell_style}">{card["song"]}</div>',
                            unsafe_allow_html=True
                        )

                st.markdown("")
                pdf_buf = build_bingo_pdf(cards, city, size=bsize)
                st.download_button(
                    label="🖨️ Download Printable PDF",
                    data=pdf_buf,
                    file_name=f"{city.replace(' ', '_')}_Bingo_{bsize}x{bsize}.pdf",
                    mime="application/pdf",
                )


# ═══════════════════════════════════════════════════════════
# TAB 2 — Top 50 · Sphere 2026 Tracker
# ═══════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="gj-section-head">Phish at the Sphere · Las Vegas 2026</div>', unsafe_allow_html=True)
    st.markdown('<div class="gj-section-sub">Top 50 most-played songs since 2008 · live Sphere setlist data · updates hourly</div>', unsafe_allow_html=True)

    with st.spinner("Loading Top 50 data & Sphere setlists..."):
        global_counter_t50, global_shows_t50 = build_top50_st()
        sphere_songs_t50, sphere_dates_t50 = fetch_sphere_songs_st()

    top50 = global_counter_t50.most_common(50)
    today_str = datetime.date.today().isoformat()
    shows_done = len([d for d in sphere_dates_t50 if d <= today_str])
    shows_left = len(sphere_dates_t50) - shows_done

    # ── Possum insight bubble ──────────────────────────────
    sphere_played_songs = [s for s in global_counter_t50 if sphere_songs_t50.get(s)]
    top10_names = [s for s, _ in top50[:10]]
    top10_at_sphere = [s for s in top10_names if sphere_songs_t50.get(s)]
    most_played_sphere = sorted(sphere_songs_t50.items(), key=lambda x: len(x[1]), reverse=True)

    insights = []
    if top10_at_sphere:
        insights.append(f"Of the Top 10 all-time songs, **{len(top10_at_sphere)}** have already been played at the Sphere: {', '.join(top10_at_sphere)}.")
    if most_played_sphere:
        song_mp, dates_mp = most_played_sphere[0]
        insights.append(f"**{song_mp}** has been played the most times this Sphere run — {len(dates_mp)} night{'s' if len(dates_mp)>1 else ''}.")
    unplayed_top10 = [s for s in top10_names if not sphere_songs_t50.get(s)]
    if unplayed_top10:
        insights.append(f"Still waiting on these Top-10 staples to appear at the Sphere: **{', '.join(unplayed_top10[:3])}**{'...' if len(unplayed_top10)>3 else ''}.")
    if shows_left > 0:
        insights.append(f"**{shows_left} show{'s' if shows_left>1 else ''}** left in the Sphere run. Still plenty of time for surprises.")
    if shows_left == 0:
        insights.append("The Sphere run is complete! What a historic stretch of shows.")
    sphere_only = [s for s, _ in top50 if s not in [x for x,_ in top50[:10]] and sphere_songs_t50.get(s)]
    if sphere_only:
        insights.append(f"Some deeper cuts showed up at the Sphere: **{', '.join(sphere_only[:3])}** — the band is digging into the catalog.")

    insight_text = random.choice(insights) if insights else "The numbers don't lie — this Sphere run is something special."

    st.markdown(f"""
    <div style="display:flex;align-items:flex-start;gap:14px;margin-bottom:20px">
        <div style="font-size:48px;line-height:1">🐀</div>
        <div style="background:#1a1a2e;border:1px solid #444;border-radius:12px;border-top-left-radius:2px;padding:14px 18px;max-width:640px">
            <div style="font-size:11px;color:#888;margin-bottom:4px;font-style:italic">Possum Insight</div>
            <div style="color:#F0E68C;font-size:14px;line-height:1.6">{insight_text}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Ask Trey section ───────────────────────────────────
    st.markdown("#### 🍩 Ask Trey — Will they play it at the Sphere?")

    with st.form("ask_trey_form", clear_on_submit=False):
        trey_q = st.text_input(
            "Ask about any song:",
            placeholder="e.g. Will you play Sand? What about Tweezer?",
            key="trey_input",
        )
        trey_submit = st.form_submit_button("Ask Trey 🎸")

    if trey_submit:
        if trey_q.strip():
            try:
                matched, answer, stats = ask_trey_st(trey_q, global_counter_t50, global_shows_t50)
                st.session_state["trey_response"] = {
                    "matched": matched,
                    "answer": answer,
                    "stats": stats,
                }
            except Exception as e:
                st.session_state["trey_response"] = {
                    "matched": None,
                    "answer": f"Something broke on my end — {type(e).__name__}: {e}",
                    "stats": {},
                }
        else:
            st.session_state["trey_response"] = {
                "matched": None,
                "answer": "Type a song name to ask Trey!",
                "stats": {},
            }

    # Render stored response (persists across reruns)
    resp = st.session_state.get("trey_response")
    if resp:
        if resp["matched"]:
            st.markdown(f"""
            <div style="display:flex;align-items:flex-start;gap:14px;margin:12px 0">
                <div style="font-size:40px;line-height:1">🍩</div>
                <div style="background:#1a1a2e;border:1px solid #444;border-radius:12px;border-top-left-radius:2px;padding:14px 18px;max-width:600px">
                    <div style="font-size:11px;color:#888;margin-bottom:6px;font-style:italic">Trey on <b style="color:#F0E68C">{resp['matched']}</b></div>
                    <div style="color:#eeeeee;font-size:14px;line-height:1.6;font-style:italic">"{resp['answer']}"</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            stats = resp["stats"] or {}
            chip_style = "display:inline-block;background:#2a2a4a;color:#7ec8e3;border:1px solid #444;border-radius:20px;padding:3px 12px;font-size:12px;margin:4px"
            chips_html = ""
            if stats.get("pct") is not None:
                chips_html += f'<span style="{chip_style}">📊 {stats["pct"]}% of shows</span>'
            if stats.get("gap") is not None:
                chips_html += f'<span style="{chip_style}">⏳ {stats["gap"]} shows since last played</span>'
            if stats.get("adj") is not None:
                chips_html += f'<span style="{chip_style}">🎯 {stats["adj"]}% gap-adjusted</span>'
            if stats.get("sphere"):
                chips_html += f'<span style="display:inline-block;background:#3a2800;color:#FFFACD;border:1px solid #7B5800;border-radius:20px;padding:3px 12px;font-size:12px;margin:4px">★ Sphere: {", ".join(stats["sphere"])}</span>'
            if chips_html:
                st.markdown(f'<div style="margin-left:58px">{chips_html}</div>', unsafe_allow_html=True)
        else:
            st.warning(resp["answer"])

    st.divider()

    # ── Top 50 table ───────────────────────────────────────
    st.markdown("#### 🎵 Top 50 Most-Played Songs (2008–Present)")
    generated_ts = datetime.datetime.now().strftime("%B %d, %Y")
    st.caption(f"Based on {global_shows_t50} shows · % = chance on any given night · ★ = played at Sphere 2026 · Updated {generated_ts}")

    TIER_FILL_T50 = {
        "green":  ("#1B4D1B", "#90EE90"),
        "orange": ("#4D2E00", "#FFB347"),
        "purple": ("#2E0050", "#CE93D8"),
    }
    SPHERE_BG_T50, SPHERE_FG_T50 = "#3a2800", "#FFFACD"

    t50_header = """
    <tr>
        <th style="background:#1A1A2E;color:#F0E68C;padding:8px;text-align:center;width:6%">Rank</th>
        <th style="background:#1A1A2E;color:#F0E68C;padding:8px;text-align:left;width:36%">Song</th>
        <th style="background:#1A1A2E;color:#F0E68C;padding:8px;text-align:center;width:14%">Times Played</th>
        <th style="background:#1A1A2E;color:#F0E68C;padding:8px;text-align:center;width:12%">% Any Night</th>
        <th style="background:#1A1A2E;color:#F0E68C;padding:8px;text-align:center;width:20%">Sphere 2026</th>
        <th style="background:#1A1A2E;color:#F0E68C;padding:8px;text-align:center;width:12%">YouTube</th>
    </tr>"""

    t50_rows = ""
    for rank, (song, count) in enumerate(top50, 1):
        pct = count / global_shows_t50 * 100
        tier_key = "green" if rank <= 10 else ("orange" if rank <= 25 else "purple")
        bg, fg = TIER_FILL_T50[tier_key]

        dates_played = sphere_songs_t50.get(song, [])
        if dates_played:
            sphere_label = "★  " + ",  ".join(d[5:] for d in sorted(dates_played))
            row_bg, row_fg = SPHERE_BG_T50, SPHERE_FG_T50
        else:
            sphere_label = ""
            row_bg, row_fg = bg, fg

        yt_query = urllib.parse.quote_plus(f"Phish {song} Sphere Las Vegas 2026")
        yt_url   = f"https://www.youtube.com/results?search_query={yt_query}"
        yt_cell  = f'<a href="{yt_url}" target="_blank" style="color:#4FC3F7;font-weight:bold;text-decoration:none">▶ Watch</a>' if dates_played else ""

        t50_rows += f"""
        <tr>
            <td style="background:{bg};color:{fg};text-align:center;padding:6px;font-size:13px">{rank}</td>
            <td style="background:{row_bg};color:{row_fg};padding:6px;font-weight:bold;font-size:13px">{song}</td>
            <td style="background:{bg};color:{fg};text-align:center;padding:6px;font-size:13px">{count}</td>
            <td style="background:{bg};color:{fg};text-align:center;padding:6px;font-size:13px">{pct:.1f}%</td>
            <td style="background:{row_bg};color:{row_fg};text-align:center;padding:6px;font-size:13px;font-weight:{'bold' if dates_played else 'normal'}">{sphere_label}</td>
            <td style="background:{row_bg};text-align:center;padding:6px;font-size:13px">{yt_cell}</td>
        </tr>"""

    t50_table = f"""
    <table style="width:100%;border-collapse:collapse;font-family:Arial">
        <thead>{t50_header}</thead>
        <tbody>{t50_rows}</tbody>
    </table>
    """
    st.markdown(t50_table, unsafe_allow_html=True)

    st.markdown("""
    <div style="font-size:11px;color:#666;margin-top:10px">
    <span style="background:#1B4D1B;color:#90EE90;padding:2px 6px;border-radius:4px">Green = Top 10</span> &nbsp;
    <span style="background:#4D2E00;color:#FFB347;padding:2px 6px;border-radius:4px">Orange = Ranks 11–25</span> &nbsp;
    <span style="background:#2E0050;color:#CE93D8;padding:2px 6px;border-radius:4px">Purple = Ranks 26–50</span> &nbsp;
    <span style="background:#3a2800;color:#FFFACD;padding:2px 6px;border-radius:4px">★ Gold = Played at Sphere 2026</span>
    </div>
    """, unsafe_allow_html=True)

    # Download Top 50 as xlsx (built in memory)
    st.markdown("")
    t50_xlsx_buf = _build_top50_xlsx_buf(top50, global_shows_t50, sphere_songs_t50)
    st.download_button(
        label="⬇️ Download Top 50 Spreadsheet (.xlsx)",
        data=t50_xlsx_buf,
        file_name="Top_50_Phish_Songs_Sphere.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_top50"
    )

    # ── Top 50 Bingo Card ──────────────────────────────────
    st.divider()
    st.markdown("#### 🎲 Top 50 Bingo Card")
    st.caption("A 5×5 bingo card drawn from the all-time Top 50. "
               "Tier-coloured by rank · ★ marks songs already played at the Sphere this run.")

    if st.button("🎲 Generate Top 50 Bingo", key="gen_t50_bingo"):
        # Sample 24 songs (center = FREE) weighted lightly by rank so top songs
        # appear more often, but ranks 26-50 still get representation.
        pool = [(s, c) for s, c in top50]
        ranks = {s: i + 1 for i, (s, _) in enumerate(pool)}

        top_tier  = [s for s, _ in pool[:10]]
        mid_tier  = [s for s, _ in pool[10:25]]
        deep_tier = [s for s, _ in pool[25:50]]

        # Pick a balanced spread: 8 top, 9 mid, 7 deep = 24 squares
        random.shuffle(top_tier); random.shuffle(mid_tier); random.shuffle(deep_tier)
        picks = top_tier[:8] + mid_tier[:9] + deep_tier[:7]
        random.shuffle(picks)

        cards = []
        for i, song in enumerate(picks):
            rank = ranks.get(song, 99)
            cat  = "setlist" if rank <= 10 else ("common" if rank <= 25 else "rare")
            cards.append({"song": song, "cat": cat, "played": bool(sphere_songs_t50.get(song))})
            if i == 11:  # insert FREE at index 12 (center of 5×5)
                cards.append({"song": "★ FREE ★", "cat": "setlist", "played": False, "free": True})
        # Ensure exactly 25 cells (in case of off-by-one when picks list shifts)
        cards = cards[:25]
        while len(cards) < 25:
            cards.append({"song": "★ FREE ★", "cat": "setlist", "played": False, "free": True})

        st.session_state["t50_bingo"] = cards

    if st.session_state.get("t50_bingo"):
        bcards_t50 = st.session_state["t50_bingo"]
        cat_styles_t50 = {
            "setlist": "background:#1B4D1B;color:#90EE90;border:1px solid #2d6b2d",
            "common":  "background:#4D2E00;color:#FFB347;border:1px solid #6e4910",
            "rare":    "background:#2E0050;color:#CE93D8;border:1px solid #4d1882",
        }
        played_overlay = ("background:linear-gradient(135deg,#3a2800 0%,#5a3d00 100%);"
                          "color:#FFFACD;border:1.5px solid #FFD54F;"
                          "box-shadow:0 0 6px rgba(255,213,79,0.35)")
        free_style = ("background:linear-gradient(135deg,#5a3d00,#3a2800);color:#FFF3B0;"
                      "border:1.5px solid #FFD54F;font-family:'Shrikhand',cursive")
        cell_style_t50 = ("padding:10px 6px;text-align:center;font-size:12px;"
                          "font-weight:600;border-radius:8px;min-height:68px;"
                          "display:flex;align-items:center;justify-content:center;word-break:break-word;")

        bcol_h = st.columns(5)
        for col, label in zip(bcol_h, ["P", "H", "I", "S", "H"]):
            col.markdown(
                f'<div style="text-align:center;font-size:26px;font-weight:700;'
                f'color:#FFF3B0;letter-spacing:0.05em;font-family:Shrikhand,cursive">{label}</div>',
                unsafe_allow_html=True
            )
        for row_i in range(5):
            cols = st.columns(5)
            for col_i, col in enumerate(cols):
                card = bcards_t50[row_i * 5 + col_i]
                if card.get("free"):
                    style = free_style
                    label = card["song"]
                elif card.get("played"):
                    style = played_overlay
                    label = "★ " + card["song"]
                else:
                    style = cat_styles_t50[card["cat"]]
                    label = card["song"]
                col.markdown(
                    f'<div style="{style};{cell_style_t50}">{label}</div>',
                    unsafe_allow_html=True
                )

        st.markdown("""
        <div style="font-size:11px;color:#888;margin-top:10px">
        <span style="background:#1B4D1B;color:#90EE90;padding:2px 6px;border-radius:4px">Top 10</span> &nbsp;
        <span style="background:#4D2E00;color:#FFB347;padding:2px 6px;border-radius:4px">11–25</span> &nbsp;
        <span style="background:#2E0050;color:#CE93D8;padding:2px 6px;border-radius:4px">26–50</span> &nbsp;
        <span style="background:#3a2800;color:#FFFACD;padding:2px 6px;border-radius:4px">★ Played at Sphere</span>
        </div>
        """, unsafe_allow_html=True)

        # PDF download — strip the extra fields the PDF builder doesn't know about
        pdf_cards = [{"song": c["song"], "cat": c["cat"]} for c in bcards_t50]
        pdf_buf_t50 = build_bingo_pdf(pdf_cards, "Top 50 All-Time")
        st.markdown("")
        st.download_button(
            label="⬇️ Download Printable Bingo PDF",
            data=pdf_buf_t50,
            file_name="Top_50_Bingo.pdf",
            mime="application/pdf",
            key="dl_t50_bingo_pdf",
        )


# ═══════════════════════════════════════════════════════════
# TAB 3 — Sphere Setlist Predictor
# ═══════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="gj-section-head">Sphere Setlist Predictor</div>', unsafe_allow_html=True)
    st.markdown('<div class="gj-section-sub">Predicts setlists for upcoming Sphere shows · '
                'excludes songs already played this run · boosts songs from the last 10–15 shows</div>',
                unsafe_allow_html=True)

    with st.spinner("Loading Sphere schedule..."):
        sphere_songs_p, sphere_dates_p = fetch_sphere_songs_st()

    today_iso  = datetime.date.today().isoformat()
    played     = sorted(d for d in sphere_dates_p if d <= today_iso)
    upcoming   = sorted(d for d in sphere_dates_p if d > today_iso)

    # Summary of run so far
    st.markdown(f"""
    <div class="gj-card gj-card-accent">
        <div style="color:#FFF3B0;font-size:14px;font-weight:600;letter-spacing:0.02em">
            Sphere Run 2026
        </div>
        <div style="color:#c8c8dc;font-size:13px;margin-top:6px;line-height:1.7">
            <b style="color:#FFF3B0">{len(played)}</b> show{'s' if len(played)!=1 else ''} played
            &nbsp;·&nbsp;
            <b style="color:#FFF3B0">{len(upcoming)}</b> upcoming
            &nbsp;·&nbsp;
            <b style="color:#FFF3B0">{len(sphere_songs_p)}</b> unique songs so far
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ─── Repeat Alert ────────────────────────────────────────
    repeats = sorted(
        [(song, dates) for song, dates in sphere_songs_p.items() if len(dates) > 1],
        key=lambda x: (-len(x[1]), x[0])
    )
    if repeats:
        repeat_rows = "".join(
            f'<div style="display:flex;justify-content:space-between;padding:4px 0;'
            f'border-bottom:1px solid rgba(255,255,255,0.06)">'
            f'<span style="color:#FFE082;font-weight:600">{song}</span>'
            f'<span style="color:#c8c8dc;font-size:12px">'
            f'{len(dates)}× — {", ".join(d[5:] for d in sorted(dates))}</span></div>'
            for song, dates in repeats
        )
        st.markdown(f"""
        <div class="gj-card" style="border-left:3px solid #FF8A65;margin-top:12px">
            <div style="color:#FF8A65;font-size:14px;font-weight:700;letter-spacing:0.04em">
                🔁 REPEAT ALERT — {len(repeats)} song{'s' if len(repeats)!=1 else ''} repeated this run
            </div>
            <div style="color:#9a9ab0;font-size:11.5px;margin:4px 0 10px 0">
                Phish normally avoids repeats within a single run. These songs have returned.
            </div>
            {repeat_rows}
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="gj-card" style="border-left:3px solid #81C784;margin-top:12px">
            <div style="color:#81C784;font-size:14px;font-weight:700;letter-spacing:0.04em">
                ✅ No repeats yet this run
            </div>
            <div style="color:#9a9ab0;font-size:11.5px;margin-top:4px">
                Every song so far has only been played once at the 2026 Sphere run.
            </div>
        </div>
        """, unsafe_allow_html=True)

    # (Last Night's Accuracy is now shown prominently at the top of the page.)

    if not upcoming:
        st.info("No remaining Sphere shows on the schedule. Check back after the next tour is announced!")
    else:
        # Pre-select from ?sphere_date=... query param if valid
        _url_sd = st.query_params.get("sphere_date", "")
        _default_idx = upcoming.index(_url_sd) if _url_sd in upcoming else 0

        target_date = st.selectbox(
            "Select an upcoming Sphere show:",
            upcoming,
            index=_default_idx,
            format_func=lambda d: datetime.date.fromisoformat(d).strftime("%A, %B %d, %Y")
        )

        if st.button("🎸 Generate Sphere Prediction", key="gen_sphere"):
            st.session_state["sphere_result"] = generate_sphere_setlist(target_date, sphere_songs_p)
            st.session_state["sphere_target"] = target_date

        if st.session_state.get("sphere_result") and st.session_state.get("sphere_target") == target_date:
            result = st.session_state["sphere_result"]
            rows = result["rows"]

            pretty_date = datetime.date.fromisoformat(target_date).strftime("%B %d, %Y")
            st.success(f"🎯 Prediction for **{pretty_date}** — {len(rows)} songs "
                       f"· based on {result['source']} · {len(result['excluded'])} excluded (already played)")

            if result["recent_dates"]:
                window_label = f"{result['recent_dates'][0]} → {result['recent_dates'][-1]}"
                st.caption(f"⚡ Recent rotation boost applied from {result['window_count']} shows "
                           f"({window_label}) — {len(result['recent_songs'])} songs flagged.")
            else:
                st.caption("⚡ No shows found in the 10–15-shows-prior window (tour hasn't run that recently).")

            # Color-coded table
            tier_colors_p = {
                "Staple":     "#FFD700",
                "Common":     "#4FC3F7",
                "Occasional": "#A5D6A7",
                "Rare":       "#CE93D8",
            }

            hdrs = ["#", "Song", "Tier", "Vegas Freq", "Global", "Gap", "Adj Score", "Role", "🔥"]
            widths = [4, 26, 11, 11, 9, 7, 10, 12, 5]
            header_html_p = "".join(
                f'<th style="background:#1A1A2E;color:#F0E68C;padding:8px;text-align:center;width:{w}%">{h}</th>'
                for h, w in zip(hdrs, widths)
            )

            body_html = ""
            prev_set_p = None
            for row in rows:
                cur_set = row.get("Set", "")
                if cur_set != prev_set_p:
                    banner = {"Set 1":"🎸 SET 1","Set 2":"🎸 SET 2","Encore":"🎤 ENCORE"}.get(cur_set, cur_set)
                    body_html += (
                        f'<tr><td colspan="9" style="background:#0d0d1e;color:#FFD54F;'
                        f'padding:10px 8px;font-weight:700;letter-spacing:0.1em;'
                        f'font-family:Shrikhand,cursive;font-size:1.05rem">{banner}</td></tr>'
                    )
                    prev_set_p = cur_set

                is_bustout = row.get("Bust Out", False)
                is_recent  = row["Recent"]
                role = row.get("Role", "")
                is_opener = role == "Opener"
                is_closer = role == "Closer"
                is_encore = cur_set == "Encore"

                if is_bustout:   bg, fg = "#1A2E1A", "#B9F6CA"
                elif is_recent:  bg, fg = "#3a1f00", "#FFCC80"
                elif is_encore:  bg, fg = "#2a1a3e", "#E1BEE7"
                elif is_closer:  bg, fg = "#2E1A1A", "#F4C2C2"
                elif is_opener:  bg, fg = "#1a2a3e", "#B3E5FC"
                else:            bg, fg = ("#1a1a2e" if row["#"] % 2 == 0 else "#16213e"), "#EEEEEE"

                tier_color = tier_colors_p.get(row["Tier"], "#FFFFFF")
                adj_color = "#66BB6A" if row["_adj"] >= 30 else ("#42A5F5" if row["_adj"] >= 25 else fg)
                recent_mark = "🔥" if is_recent else ""
                song_label = f"{row['Song']} ⭐ <span style='color:#B9F6CA;font-size:11px'>BUST OUT</span>" if is_bustout else row['Song']
                role_label = role if role else "—"

                body_html += f"""
                <tr style="background:{bg};color:{fg}">
                    <td style="text-align:center;padding:6px">{row['#']}</td>
                    <td style="padding:6px;font-weight:bold">{song_label}</td>
                    <td style="text-align:center;padding:6px;color:{tier_color}">{row['Tier']}</td>
                    <td style="text-align:center;padding:6px">{row['Vegas/Global Freq']}</td>
                    <td style="text-align:center;padding:6px">{row['Global Freq']}</td>
                    <td style="text-align:center;padding:6px">{row['Shows Since Last Played']}</td>
                    <td style="text-align:center;padding:6px;color:{adj_color};font-weight:bold">{row['Adj Score']}</td>
                    <td style="text-align:center;padding:6px">{role_label}</td>
                    <td style="text-align:center;padding:6px;font-size:16px">{recent_mark}</td>
                </tr>"""

            table_p = f"""
            <table style="width:100%;border-collapse:collapse;font-family:Arial;font-size:13px">
                <thead><tr>{header_html_p}</tr></thead>
                <tbody>{body_html}</tbody>
            </table>
            """
            st.markdown(table_p, unsafe_allow_html=True)

            st.markdown("""
            <div style="font-size:11px;color:#666;margin-top:8px">
            🔥 = recent rotation (last 10–15 shows) &nbsp;|&nbsp;
            <span style="background:#1a2a3e;color:#B3E5FC;padding:1px 4px">Blue = opener</span> &nbsp;|&nbsp;
            <span style="background:#2E1A1A;color:#F4C2C2;padding:1px 4px">Red = closer</span> &nbsp;|&nbsp;
            <span style="background:#2a1a3e;color:#E1BEE7;padding:1px 4px">Purple = encore</span> &nbsp;|&nbsp;
            <span style="background:#1A2E1A;color:#B9F6CA;padding:1px 4px">Green ⭐ = Bust Out (gap > 500)</span>
            </div>
            """, unsafe_allow_html=True)

            # Highlights
            st.divider()
            top_p = max(rows, key=lambda r: r["_adj"])
            bustouts_p = [r for r in rows if r.get("Bust Out")]
            s2_closer_p = next((r for r in rows if r.get("Set")=="Set 2" and r.get("Role")=="Closer"), None)
            encores_p = [r for r in rows if r.get("Set") == "Encore"]
            recent_hits = [r for r in rows if r["Recent"]]

            st.markdown(f"**🎯 Top pick:** {top_p['Song']} ({top_p['Adj Score']} adj score)")
            if recent_hits:
                st.markdown(f"**🔥 Current rotation hits:** {', '.join(r['Song'] for r in recent_hits[:6])}"
                            f"{'...' if len(recent_hits)>6 else ''}")
            if bustouts_p:
                b = bustouts_p[0]
                st.markdown(f"**⭐ Bust Out:** {b['Song']} — overdue by {b['_gap']} shows.")
            if s2_closer_p:
                st.markdown(f"**🎬 Set 2 closer:** {s2_closer_p['Song']}")
            if encores_p:
                st.markdown(f"**🎤 Encore:** {' → '.join(e['Song'] for e in encores_p)}")

            # Share link
            render_share_box({"sphere_date": target_date}, key="sphere")

            # Excluded list (already played at Sphere)
            if result["excluded"]:
                with st.expander(f"🚫 Excluded — {len(result['excluded'])} songs already played at Sphere 2026"):
                    st.write(", ".join(result["excluded"]))

            # Download xlsx
            st.divider()
            sphere_xlsx = build_xlsx(
                [{**r, "City Freq": r["Vegas/Global Freq"]} for r in rows],
                f"Sphere {pretty_date}",
                result["city_shows"],
            )
            st.download_button(
                label="⬇️ Download Sphere Prediction (.xlsx)",
                data=sphere_xlsx,
                file_name=f"Sphere_{target_date}_Prediction.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_sphere_xlsx",
            )

            # ── Sphere Bingo ───────────────────────────────────────
            st.divider()
            st.markdown("#### 🎲 Sphere Bingo Card")
            st.caption(f"Bingo card based on the prediction for {pretty_date} · "
                       "🟡 top picks · 🔵 globally common · 🟣 rare / uncommon · ★ FREE center")
            sphere_bsize_label = st.radio(
                "Card size", ["5×5 Classic", "3×3 Quick"],
                horizontal=True, key="sphere_bingo_size",
            )
            sphere_bsize = 3 if sphere_bsize_label.startswith("3") else 5

            if st.button("🎲 Generate Sphere Bingo", key="gen_sphere_bingo"):
                global_counter_b, global_shows_b, _, current_gap_b, total_shows_b, _ = load_data()

                def gscore(song):
                    g = current_gap_b.get(song, total_shows_b)
                    return (global_counter_b[song] / global_shows_b) * (1 + math.log(g+1) / math.log(total_shows_b+1))

                # Top 10 from the prediction itself
                setlist_picks = [r["Song"] for r in sorted(rows, key=lambda r: r["_adj"], reverse=True)[:10]]
                used = set(setlist_picks) | set(result["excluded"])

                common_pool = sorted(
                    [(s, gscore(s)) for s, c in global_counter_b.items()
                     if 5 <= (c / global_shows_b) * 100 < 15 and s not in used],
                    key=lambda x: x[1], reverse=True
                )
                common_picks = [s for s, _ in common_pool[:10]]
                used |= set(common_picks)

                occasional_pool = sorted(
                    [(s, gscore(s)) for s, c in global_counter_b.items()
                     if 1 <= (c / global_shows_b) * 100 < 5 and s not in used],
                    key=lambda x: x[1], reverse=True
                )
                rare_pool = sorted(
                    [(s, gscore(s)) for s, c in global_counter_b.items()
                     if (c / global_shows_b) * 100 < 1 and s not in used],
                    key=lambda x: x[1], reverse=True
                )
                rare_picks = [s for s, _ in occasional_pool[:3]] + [s for s, _ in rare_pool[:2]]

                all_bingo = setlist_picks + common_picks + rare_picks
                random.shuffle(all_bingo)
                st_set, cm_set = set(setlist_picks), set(common_picks)
                bingo_cards = [{"song": s,
                                "cat": "setlist" if s in st_set else "common" if s in cm_set else "rare"}
                               for s in all_bingo]
                bingo_cards = _inject_free_center(bingo_cards, sphere_bsize)
                st.session_state["sphere_bingo"] = bingo_cards
                st.session_state["sphere_bingo_date"] = target_date
                st.session_state["sphere_bingo_used_size"] = sphere_bsize

            if (st.session_state.get("sphere_bingo")
                    and st.session_state.get("sphere_bingo_date") == target_date
                    and st.session_state.get("sphere_bingo_used_size") == sphere_bsize):
                bcards = st.session_state["sphere_bingo"]
                cat_styles = {
                    "setlist": "background:#3a3a10;color:#FFF3B0;border:1px solid #5a5a22",
                    "common":  "background:#0d2a45;color:#8fd8f0;border:1px solid #1f5280",
                    "rare":    "background:#35104a;color:#d4a8e0;border:1px solid #6a3588",
                }
                free_style = ("background:linear-gradient(135deg,#5a3d00,#3a2800);color:#FFF3B0;"
                              "border:1.5px solid #FFD54F;font-family:'Shrikhand',cursive")
                cell_style = ("padding:10px 6px;text-align:center;font-size:12px;"
                              "font-weight:600;border-radius:8px;min-height:68px;"
                              "display:flex;align-items:center;justify-content:center;word-break:break-word;")

                bcol_labels = ["P", "H", "I"] if sphere_bsize == 3 else ["P", "H", "I", "S", "H"]
                bcol_headers = st.columns(sphere_bsize)
                for col, label in zip(bcol_headers, bcol_labels):
                    col.markdown(
                        f'<div style="text-align:center;font-size:26px;font-weight:700;'
                        f'color:#FFF3B0;letter-spacing:0.05em;font-family:Shrikhand,cursive">{label}</div>',
                        unsafe_allow_html=True
                    )
                for row_i in range(sphere_bsize):
                    cols = st.columns(sphere_bsize)
                    for col_i, col in enumerate(cols):
                        card = bcards[row_i * sphere_bsize + col_i]
                        style = free_style if card.get("free") else cat_styles[card["cat"]]
                        col.markdown(
                            f'<div style="{style};{cell_style}">{card["song"]}</div>',
                            unsafe_allow_html=True
                        )

                pdf_buf = build_bingo_pdf(bcards, f"Sphere {pretty_date}", size=sphere_bsize)
                st.markdown("")
                st.download_button(
                    label="⬇️ Download Printable Bingo PDF",
                    data=pdf_buf,
                    file_name=f"Sphere_{target_date}_Bingo_{sphere_bsize}x{sphere_bsize}.pdf",
                    mime="application/pdf",
                    key="dl_sphere_bingo_pdf",
                )


# ═══════════════════════════════════════════════════════════
# Methodology footer (global — shown under every tab)
# ═══════════════════════════════════════════════════════════
st.markdown("<div style='margin-top:40px'></div>", unsafe_allow_html=True)
render_methodology_footer()
